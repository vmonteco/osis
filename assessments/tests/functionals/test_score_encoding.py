import datetime
import time
from urllib import request

import faker
import functools
import magic
import os
import parse
import pendulum
import pyvirtualdisplay
import random
import shutil
import tempfile
from django.conf import settings
from django.contrib.auth.models import Group, Permission
from django.contrib.staticfiles.testing import StaticLiveServerTestCase
from django.test import tag
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select

from attribution.tests.factories.attribution import AttributionFactory
from base.tests.factories.academic_calendar import (AcademicCalendarExamSubmissionFactory,
                                                    AcademicCalendarFactory)
from base.tests.factories.academic_year import AcademicYearFactory
from base.tests.factories.exam_enrollment import ExamEnrollmentFactory
from base.tests.factories.learning_unit import LearningUnitFactory
from base.tests.factories.learning_unit_enrollment import \
    LearningUnitEnrollmentFactory
from base.tests.factories.learning_unit_year import LearningUnitYearFactory
from base.tests.factories.offer_enrollment import OfferEnrollmentFactory
from base.tests.factories.offer_year import OfferYearFactory
from base.tests.factories.offer_year_calendar import OfferYearCalendarFactory
from base.tests.factories.person import PersonFactory
from base.tests.factories.program_manager import ProgramManagerFactory
from base.tests.factories.session_exam_calendar import \
    SessionExamCalendarFactory
from base.tests.factories.session_examen import SessionExamFactory
from base.tests.factories.student import StudentFactory
from base.tests.factories.tutor import TutorFactory
from base.tests.factories.user import SuperUserFactory, UserFactory


class BusinessMixin:
    def create_user(self):
        return UserFactory()

    def create_super_user(self):
        return SuperUserFactory()

    def add_group(self, user, *group_names):
        for name in group_names:
            group, created = Group.objects.get_or_create(name=name)
            group.user_set.add(user)

    def add_permission(self, user, *permission_names):
        for permission_name in permission_names:
            if '.' in permission_name:
                label, codename = permission_name.split('.')
                permission = Permission.objects.get(codename=codename, content_type__app_label=label)
            else:
                permission = Permission.objects.get(codename=permission_name)

            user.user_permissions.add(permission)


@tag('selenium')
class SeleniumTestCase(StaticLiveServerTestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.sel_settings = settings.SELENIUM_SETTINGS
        print("### Virtual Display : {}".format(cls.sel_settings.get('VIRTUAL_DISPLAY')))
        cls.screen_size = (cls.sel_settings.get('SCREEN_WIDTH'), cls.sel_settings.get('SCREEN_HIGH'))
        cls.full_path_temp_dir = tempfile.mkdtemp('osis-selenium')
        if cls.sel_settings.get('VIRTUAL_DISPLAY'):
            cls.display = pyvirtualdisplay.Display(visible=0, size=cls.screen_size)
            cls.display.start()

        if cls.sel_settings.get('WEB_BROWSER').upper() == 'FIREFOX':
            fp = webdriver.FirefoxProfile()
            fp.set_preference('browser.download.dir', cls.full_path_temp_dir)
            fp.set_preference('browser.download.folderList', 2)
            fp.set_preference('browser.download.manager.showWhenStarting', False)
            fp.set_preference('pdfjs.disabled', True)
            known_mimes = ['application/vnd.ms-excel',
                           'application/pdf',
                           'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']
            fp.set_preference('browser.helperApps.neverAsk.saveToDisk', ','.join(known_mimes))
            cls.driver = webdriver.Firefox(executable_path=cls.sel_settings.get('GECKO_DRIVER'),
                                           firefox_profile=fp)

        if cls.sel_settings.get('WEB_BROWSER').upper() == 'CHROME':
            options = webdriver.ChromeOptions()
            options.add_experimental_option('prefs', {
                 'download.default_directory': cls.full_path_temp_dir,
                 'download.prompt_for_download': False,
                 'download.directory_upgrade': True,
                 'safebrowsing.enabled': True
             })
            cls.driver = webdriver.Chrome(chrome_options=options)

        cls.driver.implicitly_wait(10)
        cls.driver.set_window_size(*cls.screen_size)

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.full_path_temp_dir)
        cls.driver.quit()
        if cls.sel_settings.get('VIRTUAL_DISPLAY'):
            cls.display.stop()

        super().tearDownClass()

    def get_url_by_name(self, url_name, *args, **kwargs):
        return request.urljoin(self.live_server_url, reverse(url_name, args=args, kwargs=kwargs))

    def goto(self, url_name, *args, **kwargs):
        url = self.get_url_by_name(url_name, *args, **kwargs)
        self.driver.get(url)

    def fill_by_id(self, field_id, value):
        field = self.driver.find_element_by_id(field_id)
        field.clear()
        field.send_keys(value)

    def login(self, username, password='password123'):
        self.goto('login')
        self.fill_by_id('id_username', username)
        self.fill_by_id('id_password', password)
        self.click_on('post_login_btn')

    def click_on(self, id_):
        self.driver.find_element_by_id(id_).click()

    def get_element(self, id_):
        return self.driver.find_element_by_id(id_)

    def get_element_text(self, id_):
        return self.get_element(id_).text

    def assertElementTextEqualInt(self, id_, value):
        self.assertEqual(int(self.get_element_text(id_)), value)

    def assertElementTextEqual(self, id_, text):
        self.assertEqual(self.get_element_text(id_), text)

    def assertCurrentUrl(self, url_name, *args, **kwargs):
        self.assertEqual(
            self.driver.current_url,
            self.get_url_by_name(url_name, *args, **kwargs)
        )


class FunctionalTest(SeleniumTestCase, BusinessMixin):
    def test_01_scenario_modifier_periode_encoding(self):
        user = self.create_super_user()
        academic_year = AcademicYearFactory(year=pendulum.today().year-1)
        academic_calendar = AcademicCalendarFactory.build(academic_year=academic_year)
        academic_calendar.save(functions=[])
        self.login(user.username)

        self.goto('academic_calendar_read', academic_calendar_id=academic_calendar.id)
        self.click_on('bt_academic_calendar_edit')

        new_date = academic_calendar.start_date - datetime.timedelta(days=5)
        new_date_str = new_date.strftime('%m/%d/%Y')

        self.fill_by_id('txt_start_date', new_date_str)

        self.driver.execute_script("scroll(0, 250)")
        self.click_on('bt_academic_calendar_save')

        self.assertCurrentUrl('academic_calendar_form',academic_calendar_id=academic_calendar.id)

        self.assertElementTextEqual('ac_start_date', new_date_str)

    def test_01_scenario_modifier_period_encoding_date_fin(self):
        user = self.create_super_user()
        academic_year = AcademicYearFactory(year=pendulum.today().year-1)
        academic_calendar = AcademicCalendarFactory.build(academic_year=academic_year)
        academic_calendar.save(functions=[])

        self.login(user.username)

        self.goto('academic_calendar_read', academic_calendar_id=academic_calendar.id)
        self.click_on('bt_academic_calendar_edit')

        new_date = academic_calendar.end_date + datetime.timedelta(days=5)
        new_date_str = new_date.strftime('%m/%d/%Y')
        self.fill_by_id('txt_end_date', new_date_str)

        self.driver.execute_script("scroll(0, 250)")
        self.click_on('bt_academic_calendar_save')

        self.assertCurrentUrl('academic_calendar_form', academic_calendar_id=academic_calendar.id)

        self.assertElementTextEqual('ac_end_date', new_date_str)


# class Scenario3FunctionalTest(SeleniumTestCase, BusinessMixin):
    def test_03(self):
        user = self.create_user()
        self.add_group(user, 'program_managers')
        self.add_permission(user, 'assessments.can_access_scoreencoding')

        start_date = timezone.now() + datetime.timedelta(days=20)

        academic_year = AcademicYearFactory(year=pendulum.today().year-1)

        academic_calendar = AcademicCalendarExamSubmissionFactory.build(
            academic_year=academic_year,
            start_date=start_date,
            end_date=start_date + datetime.timedelta(days=10),
        )
        academic_calendar.save(functions=[])

        person = PersonFactory(user=user, language='fr-be')
        offer_year = OfferYearFactory(academic_year=academic_year)

        ProgramManagerFactory(offer_year=offer_year, person=person)

        sec = SessionExamCalendarFactory(academic_calendar=academic_calendar)

        self.login(user.username)

        self.goto('scores_encoding')

        warning_messages = self.driver.find_element_by_id('pnl_warning_messages')

        result = parse.parse(
            "La période d'encodage des notes pour la session {session:d} sera ouverte à partir du {date}",
            warning_messages.text
        )

        self.assertIsNotNone(result)

        self.assertEqual(result['date'], academic_calendar.start_date.strftime('%d/%m/%Y'))


# class Scenario4FunctionalTest(SeleniumTestCase, BusinessMixin):
    def test_04(self):
        user = self.create_user()
        self.add_group(user, 'program_managers')
        self.add_permission(user, 'can_access_academic_calendar', 'assessments.can_access_scoreencoding')

        academic_year = AcademicYearFactory(year=pendulum.today().year-1)
        academic_calendar = AcademicCalendarExamSubmissionFactory.build(academic_year=academic_year)
        academic_calendar.save(functions=[])

        person = PersonFactory(user=user, language='fr-be')
        offer_year_factory = functools.partial(OfferYearFactory, academic_year=academic_year)

        acronyms = ['PHYS11BA', 'ECON2M1', 'PHYS1BA', 'PHYS2M1', 'PHYS2MA']
        offers = {
            acronym: offer_year_factory(acronym=acronym)
            for acronym in acronyms
        }

        program_manager_factory = functools.partial(
            ProgramManagerFactory, person=person
        )

        for acronym, offer_year in offers.items():
            program_manager_factory(offer_year=offer_year)

        student1 = StudentFactory()
        student2 = StudentFactory()
        student3 = StudentFactory()

        student10 = StudentFactory()
        student11 = StudentFactory()
        student12 = StudentFactory()
        student13 = StudentFactory()
        student14 = StudentFactory()
        student15 = StudentFactory()
        student16 = StudentFactory()

        offer_enrollment1 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student1)
        offer_enrollment2 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student2)
        offer_enrollment3 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student3)

        offer_enrollment10 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student10)
        offer_enrollment11 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student11)
        offer_enrollment12 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student12)
        offer_enrollment13 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student13)
        offer_enrollment14 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student14)
        offer_enrollment15 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student15)
        offer_enrollment16 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student16)

        offer_enrollment4 = OfferEnrollmentFactory(offer_year=offers['ECON2M1'], student=student1)
        offer_enrollment5 = OfferEnrollmentFactory(offer_year=offers['ECON2M1'], student=student2)

        # unité d'enseignement = learning_unit_year
        learning_unit = LearningUnitFactory()
        learning_unit_year_1 = LearningUnitYearFactory(academic_year=academic_year, learning_unit=learning_unit)

        learning_unit2 = LearningUnitFactory()
        learning_unit_year_2 = LearningUnitYearFactory(academic_year=academic_year, learning_unit=learning_unit2)

        learning_unit3 = LearningUnitFactory()
        learning_unit_year_3 = LearningUnitYearFactory(academic_year=academic_year, learning_unit=learning_unit3)

        learning_unit_enrollment1 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment1, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment2 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment2, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment3 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment3, learning_unit_year=learning_unit_year_1)

        learning_unit_enrollment10 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment10, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment11 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment11, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment12 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment12, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment13 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment13, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment14 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment14, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment15 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment15, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment16 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment16, learning_unit_year=learning_unit_year_1)

        learning_unit_enrollment4 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment4, learning_unit_year=learning_unit_year_2)
        learning_unit_enrollment5 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment5, learning_unit_year=learning_unit_year_2)

        learning_unit_enrollment6 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment1, learning_unit_year=learning_unit_year_3)
        learning_unit_enrollment7 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment5, learning_unit_year=learning_unit_year_3)

        session_exam_calendar = SessionExamCalendarFactory(academic_calendar=academic_calendar)

        session_exam_phys11ba = SessionExamFactory(learning_unit_year=learning_unit_year_1, number_session=session_exam_calendar.number_session, offer_year=offers['PHYS11BA'])

        session_exam_econ2m1 = SessionExamFactory(learning_unit_year=learning_unit_year_2, number_session=session_exam_calendar.number_session, offer_year=offers['ECON2M1'])

        session_exam_3 = SessionExamFactory(learning_unit_year=learning_unit_year_3, number_session=session_exam_calendar.number_session, offer_year=offers['ECON2M1'])
        session_exam_4 = SessionExamFactory(learning_unit_year=learning_unit_year_3, number_session=session_exam_calendar.number_session, offer_year=offers['PHYS11BA'])

        offer_year_calendar = OfferYearCalendarFactory(academic_calendar=academic_calendar, offer_year=offers['PHYS11BA'])
        offer_year_calendar = OfferYearCalendarFactory(academic_calendar=academic_calendar, offer_year=offers['ECON2M1'])

        ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment1, session_exam=session_exam_phys11ba)
        ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment2, session_exam=session_exam_phys11ba)
        ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment3, session_exam=session_exam_phys11ba)

        exam_enrollment_10 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment10, session_exam=session_exam_phys11ba)
        exam_enrollment_11 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment11, session_exam=session_exam_phys11ba)
        exam_enrollment_12 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment12, session_exam=session_exam_phys11ba)
        exam_enrollment_13 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment13, session_exam=session_exam_phys11ba)
        exam_enrollment_14 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment14, session_exam=session_exam_phys11ba)
        exam_enrollment_15 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment15, session_exam=session_exam_phys11ba)
        exam_enrollment_16 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment16, session_exam=session_exam_phys11ba)


        ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment4, session_exam=session_exam_econ2m1)
        ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment5, session_exam=session_exam_econ2m1)

        ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment6, session_exam=session_exam_3)
        ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment7, session_exam=session_exam_4)

        self.login(user.username)

        self.goto('scores_encoding')

        select = Select(self.get_element('slt_offer_list_selection'))

        all_options = set(option.text for option in select.options)

        all_offers = set(offers.keys())

        self.assertSetEqual({'Tout'}, all_options - all_offers)

        self.assertElementTextEqualInt('scores_encoding_learning_units', 3)
        self.assertEqual(len(all_options - {'Tout'}), 5)

        self.click_on('lnk_encode_{}'.format(learning_unit_year_1.id))

        # progression = self.driver.find_element_by_id('luy_progression').text
        # self.assertEqual(progression, '0 / 10')

        self.assertFalse(learning_unit_year_1.decimal_scores)
        self.assertElementTextEqual(
            'message_decimal_accepted',
            "Les notes de ce cours ne peuvent PAS recevoir de valeurs décimales."
        )

        self.assertElementTextEqualInt('number_of_enrollments', 10)

        the_first = 1
        element = self.driver.find_element_by_css_selector("[tabindex='%d']" % the_first)
        element_id = element.get_attribute('id')

        element.clear()
        element.send_keys(12)

        enrollment_id = int(element_id.split('_')[-1])

        self.click_on('bt_save_online_encoding_up')
        self.assertElementTextEqual('luy_progression', '1 / 10')

        self.assertElementTextEqualInt('enrollment_note_{}'.format(enrollment_id), 12)

        element = self.driver.find_element_by_css_selector('td#enrollment_status_{} span'.format(enrollment_id))
        self.assertIn('glyphicon-send', element.get_attribute('class').split())

        self.click_on('lnk_encode')

        note_enrollments = {}

        for counter in range(2, 11):
            element = self.driver.find_element_by_css_selector("[tabindex='%d']" % counter)
            element_id = element.get_attribute('id')
            enrollment_id = int(element_id.split('_')[-1])
            self.fill_by_id(element_id, counter)
            note_enrollments[enrollment_id] = counter

        self.click_on('bt_save_online_encoding_up')

        self.assertElementTextEqual('luy_progression', '10 / 10')

        for enrollment_id, value in note_enrollments.items():
            self.assertElementTextEqualInt('enrollment_note_{}'.format(enrollment_id), value)

        self.click_on('lnk_encode')
        note_enrollments = set()

        for counter in range(1, 11):
            element = self.driver.find_element_by_css_selector("[tabindex='%d']" % counter)
            element_id = element.get_attribute('id')
            enrollment_id = int(element_id.split('_')[-1])
            element.clear()
            note_enrollments.add(enrollment_id)

        self.click_on('bt_save_online_encoding_up')

        self.assertElementTextEqual('luy_progression', '0 / 10')

        for enrollment_id in note_enrollments:
            self.assertElementTextEqual('enrollment_note_{}'.format(enrollment_id), '-')


# class Scenario5FunctionalTest(SeleniumTestCase, BusinessMixin):
    def test_05(self):
        user = self.create_user()
        self.add_group(user, 'program_managers', 'tutors')
        self.add_permission(user, 'can_access_academic_calendar', 'assessments.can_access_scoreencoding')

        academic_year = AcademicYearFactory(year=pendulum.today().year-1)
        academic_calendar = AcademicCalendarExamSubmissionFactory.build(academic_year=academic_year)
        academic_calendar.save(functions=[])

        person = PersonFactory(
            user=user,
            first_name=user.first_name,
            last_name=user.last_name,
            language='fr-be'
        )

        offer_year_factory = functools.partial(OfferYearFactory, academic_year=academic_year)

        acronyms = ['PHYS11BA', 'ECON2M1', 'PHYS1BA', 'PHYS2M1', 'PHYS2MA']
        offers = {
            acronym: offer_year_factory(acronym=acronym)
            for acronym in acronyms
        }

        program_manager_factory = functools.partial(
            ProgramManagerFactory, person=person
        )

        for acronym, offer_year in offers.items():
            program_manager_factory(offer_year=offer_year)

        student1 = StudentFactory()
        student2 = StudentFactory()
        student3 = StudentFactory()

        student10 = StudentFactory()
        student11 = StudentFactory()
        student12 = StudentFactory()
        student13 = StudentFactory()
        student14 = StudentFactory()
        student15 = StudentFactory()
        student16 = StudentFactory()

        offer_enrollment1 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student1)
        offer_enrollment2 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student2)
        offer_enrollment3 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student3)

        offer_enrollment10 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student10)
        offer_enrollment11 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student11)
        offer_enrollment12 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student12)
        offer_enrollment13 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student13)
        offer_enrollment14 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student14)
        offer_enrollment15 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student15)
        offer_enrollment16 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student16)

        offer_enrollment4 = OfferEnrollmentFactory(offer_year=offers['ECON2M1'], student=student1)
        offer_enrollment5 = OfferEnrollmentFactory(offer_year=offers['ECON2M1'], student=student2)

        # unité d'enseignement = learning_unit_year
        learning_unit = LearningUnitFactory()
        learning_unit_year_1 = LearningUnitYearFactory(academic_year=academic_year, learning_unit=learning_unit)
        tutor = TutorFactory(person=person)

        attribution = AttributionFactory(
            tutor=tutor, learning_unit_year=learning_unit_year_1,
            score_responsible=True
        )

        learning_unit2 = LearningUnitFactory()
        learning_unit_year_2 = LearningUnitYearFactory(academic_year=academic_year, learning_unit=learning_unit2)

        learning_unit3 = LearningUnitFactory()
        learning_unit_year_3 = LearningUnitYearFactory(academic_year=academic_year, learning_unit=learning_unit3)

        learning_unit_enrollment1 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment1, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment2 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment2, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment3 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment3, learning_unit_year=learning_unit_year_1)

        learning_unit_enrollment10 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment10, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment11 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment11, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment12 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment12, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment13 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment13, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment14 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment14, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment15 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment15, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment16 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment16, learning_unit_year=learning_unit_year_1)

        learning_unit_enrollment4 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment4, learning_unit_year=learning_unit_year_2)
        learning_unit_enrollment5 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment5, learning_unit_year=learning_unit_year_2)

        learning_unit_enrollment6 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment1, learning_unit_year=learning_unit_year_3)
        learning_unit_enrollment7 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment5, learning_unit_year=learning_unit_year_3)

        session_exam_calendar = SessionExamCalendarFactory(academic_calendar=academic_calendar)

        session_exam_phys11ba = SessionExamFactory(learning_unit_year=learning_unit_year_1, number_session=session_exam_calendar.number_session, offer_year=offers['PHYS11BA'])
        session_exam_econ2m1 = SessionExamFactory(learning_unit_year=learning_unit_year_2, number_session=session_exam_calendar.number_session, offer_year=offers['ECON2M1'])

        session_exam_3 = SessionExamFactory(learning_unit_year=learning_unit_year_3, number_session=session_exam_calendar.number_session, offer_year=offers['ECON2M1'])
        session_exam_4 = SessionExamFactory(learning_unit_year=learning_unit_year_3, number_session=session_exam_calendar.number_session, offer_year=offers['PHYS11BA'])

        offer_year_calendar = OfferYearCalendarFactory(academic_calendar=academic_calendar, offer_year=offers['PHYS11BA'])
        offer_year_calendar = OfferYearCalendarFactory(academic_calendar=academic_calendar, offer_year=offers['ECON2M1'])

        exam_enrollment_1 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment1, session_exam=session_exam_phys11ba)
        exam_enrollment_2 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment2, session_exam=session_exam_phys11ba)
        exam_enrollment_3 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment3, session_exam=session_exam_phys11ba)

        exam_enrollment_10 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment10, session_exam=session_exam_phys11ba)
        exam_enrollment_11 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment11, session_exam=session_exam_phys11ba)
        exam_enrollment_12 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment12, session_exam=session_exam_phys11ba)
        exam_enrollment_13 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment13, session_exam=session_exam_phys11ba)
        exam_enrollment_14 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment14, session_exam=session_exam_phys11ba)
        exam_enrollment_15 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment15, session_exam=session_exam_phys11ba)
        exam_enrollment_16 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment16, session_exam=session_exam_phys11ba)


        ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment4, session_exam=session_exam_econ2m1)
        ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment5, session_exam=session_exam_econ2m1)

        ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment6, session_exam=session_exam_3)
        ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment7, session_exam=session_exam_4)

        self.login(user.username)

        self.goto('scores_encoding')

        select = Select(self.driver.find_element_by_id('slt_offer_list_selection'))
        select.select_by_visible_text('PHYS11BA')
        self.click_on('bt_submit_offer_search')
        time.sleep(1)
        self.click_on('lnk_via_excel')
        time.sleep(1)
        self.click_on('lnk_scores_encoding_download_{}'.format(learning_unit_year_1.id))
        time.sleep(1)
        filename = 'session_{}_{}_{}.xlsx'.format(academic_year.year,
                                                  session_exam_calendar.number_session,
                                                  learning_unit_year_1.acronym)
        full_path = os.path.join(self.full_path_temp_dir, filename)

        self.assertTrue(os.path.exists(full_path))

        exam_enrollments = [
            exam_enrollment_1, exam_enrollment_2, exam_enrollment_3,
            exam_enrollment_10, exam_enrollment_11, exam_enrollment_12,
            exam_enrollment_13, exam_enrollment_14, exam_enrollment_15,
            exam_enrollment_16
        ]

        updated_values = self.update_xlsx(full_path, exam_enrollments)

        self.goto('online_encoding', learning_unit_year_id=learning_unit_year_1.id)
        self.driver.save_screenshot(os.path.join(self.full_path_temp_dir, 'scenario5-before_xls.png'))

        self.click_on('bt_upload_score_modal')
        time.sleep(1)
        self.driver.execute_script("document.getElementById('fle_scores_input_file').style.display = 'block'")
        self.fill_by_id('fle_scores_input_file', full_path)
        time.sleep(1)

        self.click_on('bt_submit_upload_score_modal')

        self.assertElementTextEqual('luy_progression', '10 / 10')
        self.driver.save_screenshot(os.path.join(self.full_path_temp_dir, 'scenario5-final.png'))

        for enrollment_id, (key, value) in updated_values.items():
            element_id = 'enrollment_{}_{}'.format(key, enrollment_id)
            value = {'T': 'Tricherie', 'A': 'Absence injustifiée'}.get(value, value)

            self.assertElementTextEqual(element_id, str(value))


    def update_xlsx(self, filename, exam_enrollments):
        fake = faker.Faker()

        wb = load_workbook(filename)

        enrollments = {}

        sheet = wb.active

        if sheet.max_row > 11:
            start_row = 12

            nomas = {
                enrollment.learning_unit_enrollment.offer_enrollment.student.registration_id: {
                    'enrollment': enrollment,
                    'position': None
                }
                for enrollment in exam_enrollments
            }

            for counter in range(len(exam_enrollments)):
                noma = sheet['E{}'.format(counter + start_row)].value
                nomas[noma]['position'] = counter

            for noma, info in nomas.items():
                left_or_right = bool(random.getrandbits(1))
                selected_column = 'H' if left_or_right else 'I'

                if left_or_right:
                    value = random.randint(0, 20)
                    key = 'note'
                else:
                    value = fake.random_element(elements=('A', 'T'))
                    key = 'justification'

                sheet['{}{}'.format(selected_column, info['position'] + start_row)] = value

                enrollments[info['enrollment'].id] = key, value

        wb.save(filename=filename)
        return enrollments


# class Scenario6FunctionalTest(SeleniumTestCase, BusinessMixin):
    def test_06(self):
        user = self.create_user()
        self.add_group(user, 'program_managers', 'tutors')
        self.add_permission(user, 'can_access_academic_calendar', 'assessments.can_access_scoreencoding')

        academic_year = AcademicYearFactory(year=pendulum.today().year-1)
        academic_calendar = AcademicCalendarExamSubmissionFactory.build(academic_year=academic_year)
        academic_calendar.save(functions=[])

        person = PersonFactory(
            user=user,
            first_name=user.first_name,
            last_name=user.last_name,
            language='fr-be'
        )

        offer_year_factory = functools.partial(OfferYearFactory, academic_year=academic_year)

        acronyms = ['PHYS11BA', 'ECON2M1', 'PHYS1BA', 'PHYS2M1', 'PHYS2MA']
        offers = {
            acronym: offer_year_factory(acronym=acronym)
            for acronym in acronyms
        }

        program_manager_factory = functools.partial(
            ProgramManagerFactory, person=person
        )

        for acronym, offer_year in offers.items():
            program_manager_factory(offer_year=offer_year)

        student1 = StudentFactory()
        student2 = StudentFactory()
        student3 = StudentFactory()

        student10 = StudentFactory()
        student11 = StudentFactory()
        student12 = StudentFactory()
        student13 = StudentFactory()
        student14 = StudentFactory()
        student15 = StudentFactory()
        student16 = StudentFactory()

        offer_enrollment1 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student1)
        offer_enrollment2 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student2)
        offer_enrollment3 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student3)

        offer_enrollment10 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student10)
        offer_enrollment11 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student11)
        offer_enrollment12 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student12)
        offer_enrollment13 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student13)
        offer_enrollment14 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student14)
        offer_enrollment15 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student15)
        offer_enrollment16 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student16)

        offer_enrollment4 = OfferEnrollmentFactory(offer_year=offers['ECON2M1'], student=student1)
        offer_enrollment5 = OfferEnrollmentFactory(offer_year=offers['ECON2M1'], student=student2)

        # unité d'enseignement = learning_unit_year
        learning_unit = LearningUnitFactory()
        learning_unit_year_1 = LearningUnitYearFactory(academic_year=academic_year, learning_unit=learning_unit)

        tutor = TutorFactory(person=person)

        attribution = AttributionFactory(
            tutor=tutor, learning_unit_year=learning_unit_year_1,
            score_responsible=True
        )

        learning_unit2 = LearningUnitFactory()
        learning_unit_year_2 = LearningUnitYearFactory(academic_year=academic_year, learning_unit=learning_unit2)

        learning_unit3 = LearningUnitFactory()
        learning_unit_year_3 = LearningUnitYearFactory(academic_year=academic_year, learning_unit=learning_unit3)

        learning_unit_enrollment1 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment1, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment2 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment2, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment3 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment3, learning_unit_year=learning_unit_year_1)

        learning_unit_enrollment10 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment10, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment11 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment11, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment12 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment12, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment13 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment13, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment14 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment14, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment15 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment15, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment16 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment16, learning_unit_year=learning_unit_year_1)

        learning_unit_enrollment4 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment4, learning_unit_year=learning_unit_year_2)
        learning_unit_enrollment5 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment5, learning_unit_year=learning_unit_year_2)

        learning_unit_enrollment6 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment1, learning_unit_year=learning_unit_year_3)
        learning_unit_enrollment7 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment5, learning_unit_year=learning_unit_year_3)

        session_exam_calendar = SessionExamCalendarFactory(academic_calendar=academic_calendar)

        session_exam_phys11ba = SessionExamFactory(learning_unit_year=learning_unit_year_1, number_session=session_exam_calendar.number_session, offer_year=offers['PHYS11BA'])
        session_exam_econ2m1 = SessionExamFactory(learning_unit_year=learning_unit_year_2, number_session=session_exam_calendar.number_session, offer_year=offers['ECON2M1'])

        session_exam_3 = SessionExamFactory(learning_unit_year=learning_unit_year_3, number_session=session_exam_calendar.number_session, offer_year=offers['ECON2M1'])
        session_exam_4 = SessionExamFactory(learning_unit_year=learning_unit_year_3, number_session=session_exam_calendar.number_session, offer_year=offers['PHYS11BA'])

        offer_year_calendar = OfferYearCalendarFactory(academic_calendar=academic_calendar, offer_year=offers['PHYS11BA'])
        offer_year_calendar = OfferYearCalendarFactory(academic_calendar=academic_calendar, offer_year=offers['ECON2M1'])

        ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment1, session_exam=session_exam_phys11ba)
        ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment2, session_exam=session_exam_phys11ba)
        ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment3, session_exam=session_exam_phys11ba)

        exam_enrollment_10 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment10, session_exam=session_exam_phys11ba)
        exam_enrollment_11 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment11, session_exam=session_exam_phys11ba)
        exam_enrollment_12 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment12, session_exam=session_exam_phys11ba)
        exam_enrollment_13 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment13, session_exam=session_exam_phys11ba)
        exam_enrollment_14 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment14, session_exam=session_exam_phys11ba)
        exam_enrollment_15 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment15, session_exam=session_exam_phys11ba)
        exam_enrollment_16 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment16, session_exam=session_exam_phys11ba)


        ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment4, session_exam=session_exam_econ2m1)
        ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment5, session_exam=session_exam_econ2m1)

        ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment6, session_exam=session_exam_3)
        ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment7, session_exam=session_exam_4)


        self.login(user.username)

        self.goto('scores_encoding')

        self.assertElementTextEqualInt('scores_encoding_learning_units', 3)
        self.assertElementTextEqualInt('scores_encoding_programs', 5)

        select = Select(self.get_element('slt_offer_list_selection'))

        all_options = set(option.text for option in select.options)

        all_offers = set(offers.keys())

        self.assertSetEqual({'Tout'}, all_options - all_offers)

        self.assertElementTextEqualInt('scores_encoding_learning_units', 3)
        self.assertEqual(len(all_options - {'Tout'}), 5)

        self.click_on('lnk_encode_{}'.format(learning_unit_year_1.id))
        self.assertElementTextEqualInt('number_of_enrollments', 10)

        note_enrollments = {}

        for counter in range(1, 11):
            element = self.driver.find_element_by_css_selector("[tabindex='%d']" % counter)
            element_id = element.get_attribute('id')
            enrollment_id = int(element_id.split('_')[-1])
            self.fill_by_id(element_id, counter)
            note_enrollments[enrollment_id] = counter

        self.click_on('bt_save_online_encoding_up')

        self.assertElementTextEqual('luy_progression', '10 / 10')

        for enrollment_id, value in note_enrollments.items():
            self.assertElementTextEqualInt('enrollment_note_{}'.format(enrollment_id), value)

        self.click_on('lnk_online_double_encoding')

        for enrollment_id, value in note_enrollments.items():
            self.fill_by_id('num_double_score_{}'.format(enrollment_id), str(value + 2))

        self.click_on('bt_compare_up')

        self.driver.execute_script("scroll(0, document.body.scrollHeight)")

        for enrollment_id in note_enrollments:
            self.click_on('bt_take_reencoded_{}'.format(enrollment_id))

        self.click_on('bt_submit_online_double_encoding_validation')

        self.driver.execute_script("scroll(0, document.body.scrollHeight)")

        for enrollment_id, value in note_enrollments.items():
            self.assertElementTextEqualInt('enrollment_note_{}'.format(enrollment_id), value + 2)

# class Scenario7FunctionalTest(SeleniumTestCase, BusinessMixin):
    def test_07(self):
        user = self.create_user()
        self.add_group(user, 'program_managers', 'tutors')
        self.add_permission(user, 'can_access_academic_calendar', 'assessments.can_access_scoreencoding')

        academic_year = AcademicYearFactory(year=pendulum.today().year-1)
        academic_calendar = AcademicCalendarExamSubmissionFactory.build(academic_year=academic_year)
        academic_calendar.save(functions=[])

        person = PersonFactory(
            user=user,
            first_name=user.first_name,
            last_name=user.last_name,
            language='fr-be'
        )

        offer_year_factory = functools.partial(OfferYearFactory, academic_year=academic_year)

        acronyms = ['PHYS11BA', 'ECON2M1', 'PHYS1BA', 'PHYS2M1', 'PHYS2MA']
        offers = {
            acronym: offer_year_factory(acronym=acronym)
            for acronym in acronyms
        }

        program_manager_factory = functools.partial(
            ProgramManagerFactory, person=person
        )

        for acronym, offer_year in offers.items():
            program_manager_factory(offer_year=offer_year)

        student1 = StudentFactory()
        student2 = StudentFactory()
        student3 = StudentFactory()

        student10 = StudentFactory()
        student11 = StudentFactory()
        student12 = StudentFactory()
        student13 = StudentFactory()
        student14 = StudentFactory()
        student15 = StudentFactory()
        student16 = StudentFactory()

        offer_enrollment1 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student1)
        offer_enrollment2 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student2)
        offer_enrollment3 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student3)

        offer_enrollment10 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student10)
        offer_enrollment11 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student11)
        offer_enrollment12 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student12)
        offer_enrollment13 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student13)
        offer_enrollment14 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student14)
        offer_enrollment15 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student15)
        offer_enrollment16 = OfferEnrollmentFactory(offer_year=offers['PHYS11BA'], student=student16)

        offer_enrollment4 = OfferEnrollmentFactory(offer_year=offers['ECON2M1'], student=student1)
        offer_enrollment5 = OfferEnrollmentFactory(offer_year=offers['ECON2M1'], student=student2)

        # unité d'enseignement = learning_unit_year
        learning_unit = LearningUnitFactory()
        learning_unit_year_1 = LearningUnitYearFactory(academic_year=academic_year, learning_unit=learning_unit)

        tutor = TutorFactory(person=person)

        attribution = AttributionFactory(
            tutor=tutor, learning_unit_year=learning_unit_year_1,
            score_responsible=True
        )

        learning_unit2 = LearningUnitFactory()
        learning_unit_year_2 = LearningUnitYearFactory(academic_year=academic_year, learning_unit=learning_unit2)

        learning_unit3 = LearningUnitFactory()
        learning_unit_year_3 = LearningUnitYearFactory(academic_year=academic_year, learning_unit=learning_unit3)

        learning_unit_enrollment1 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment1, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment2 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment2, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment3 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment3, learning_unit_year=learning_unit_year_1)

        learning_unit_enrollment10 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment10, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment11 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment11, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment12 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment12, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment13 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment13, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment14 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment14, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment15 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment15, learning_unit_year=learning_unit_year_1)
        learning_unit_enrollment16 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment16, learning_unit_year=learning_unit_year_1)

        learning_unit_enrollment4 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment4, learning_unit_year=learning_unit_year_2)
        learning_unit_enrollment5 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment5, learning_unit_year=learning_unit_year_2)

        learning_unit_enrollment6 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment1, learning_unit_year=learning_unit_year_3)
        learning_unit_enrollment7 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment5, learning_unit_year=learning_unit_year_3)

        session_exam_calendar = SessionExamCalendarFactory(academic_calendar=academic_calendar)

        session_exam_phys11ba = SessionExamFactory(learning_unit_year=learning_unit_year_1, number_session=session_exam_calendar.number_session, offer_year=offers['PHYS11BA'])
        session_exam_econ2m1 = SessionExamFactory(learning_unit_year=learning_unit_year_2, number_session=session_exam_calendar.number_session, offer_year=offers['ECON2M1'])

        session_exam_3 = SessionExamFactory(learning_unit_year=learning_unit_year_3, number_session=session_exam_calendar.number_session, offer_year=offers['ECON2M1'])
        session_exam_4 = SessionExamFactory(learning_unit_year=learning_unit_year_3, number_session=session_exam_calendar.number_session, offer_year=offers['PHYS11BA'])

        offer_year_calendar = OfferYearCalendarFactory(academic_calendar=academic_calendar, offer_year=offers['PHYS11BA'])
        offer_year_calendar = OfferYearCalendarFactory(academic_calendar=academic_calendar, offer_year=offers['ECON2M1'])

        ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment1, session_exam=session_exam_phys11ba)
        ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment2, session_exam=session_exam_phys11ba)
        ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment3, session_exam=session_exam_phys11ba)

        exam_enrollment_10 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment10, session_exam=session_exam_phys11ba)
        exam_enrollment_11 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment11, session_exam=session_exam_phys11ba)
        exam_enrollment_12 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment12, session_exam=session_exam_phys11ba)
        exam_enrollment_13 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment13, session_exam=session_exam_phys11ba)
        exam_enrollment_14 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment14, session_exam=session_exam_phys11ba)
        exam_enrollment_15 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment15, session_exam=session_exam_phys11ba)
        exam_enrollment_16 = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment16, session_exam=session_exam_phys11ba)


        ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment4, session_exam=session_exam_econ2m1)
        ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment5, session_exam=session_exam_econ2m1)

        ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment6, session_exam=session_exam_3)
        ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment7, session_exam=session_exam_4)

        self.login(user.username)

        self.goto('scores_encoding')

        self.click_on('lnk_encode_{}'.format(learning_unit_year_1.id))
        self.assertElementTextEqualInt('number_of_enrollments', 10)

        note_enrollments = {}

        for counter in range(1, 11):
            element = self.driver.find_element_by_css_selector("[tabindex='%d']" % counter)
            element_id = element.get_attribute('id')
            enrollment_id = int(element_id.split('_')[-1])
            self.fill_by_id(element_id, counter)
            note_enrollments[enrollment_id] = counter

        self.click_on('bt_save_online_encoding_up')

        self.goto('scores_encoding')

        self.fill_by_id('txt_acronym', learning_unit_year_1.acronym)
        self.click_on('bt_submit_offer_search')
        time.sleep(1)

        self.click_on('lnk_via_paper')
        time.sleep(1)

        self.click_on('lnk_notes_printing_{}'.format(learning_unit_year_1.id))
        time.sleep(1)
        filename = 'Feuille de notes.pdf'
        full_path = os.path.join(self.full_path_temp_dir, filename)

        self.assertTrue(os.path.exists(full_path))

        mimetype = magic.from_file(full_path, mime=True)
        self.assertEqual(mimetype, 'application/pdf')

class Scenario7FunctionalTest(SeleniumTestCase, BusinessMixin):
    def test(self):
        user, person = self.create_user_person()

        academic_year, academic_calendar = self.create_academic_year_calendar()

        acronyms = ['PHYS11BA', 'ECON2M1', 'PHYS1BA', 'PHYS2M1', 'PHYS2MA']

        offers = self.create_offers(academic_year, acronyms, person=person)
        offer_year = offers['PHYS11BA']

        OfferYearCalendarFactory(academic_calendar=academic_calendar, offer_year=offer_year)

        # unité d'enseignement = learning_unit_year
        learning_unit_year = LearningUnitYearFactory(academic_year=academic_year)

        AttributionFactory(
            tutor=TutorFactory(person=person),
            learning_unit_year=learning_unit_year,
            score_responsible=True
        )

        session_exam_calendar = SessionExamCalendarFactory(academic_calendar=academic_calendar)
        session_exam = self.create_session_exam(learning_unit_year, session_exam_calendar, offer_year)
        exam_enrollments = self.create_exam_enrollments(offer_year, learning_unit_year, session_exam)

        self.login(user.username)

        self.goto('scores_encoding')

        self.click_on('lnk_encode_{}'.format(learning_unit_year.id))
        self.assertElementTextEqualInt('number_of_enrollments', len(exam_enrollments))

        note_enrollments = {}

        for counter in range(1, 11):
            element = self.driver.find_element_by_css_selector("[tabindex='%d']" % counter)
            element_id = element.get_attribute('id')
            enrollment_id = int(element_id.split('_')[-1])
            self.fill_by_id(element_id, counter)
            note_enrollments[enrollment_id] = counter

        self.click_on('bt_save_online_encoding_up')


        score_encoding = ScoresEncodingPage(self.driver, base_url=self.get_url_by_name('scores_encoding')).open()
        time.sleep(1)

        score_encoding.via_paper.click()
        # self.click_on('lnk_via_paper')
        time.sleep(1)

        self.click_on('lnk_notes_printing_{}'.format(learning_unit_year.id))
        time.sleep(1)

        filename = 'Feuille de notes.pdf'
        self.assertBrowserFileExists(filename, 'application/pdf')

    def assertBrowserFileExists(self, filename, mimetype=None):
        path = os.path.join(self.full_path_temp_dir, filename)
        self.assertTrue(os.path.exists(path))

        if mimetype:
            self.assertEqual(mimetype, magic.from_file(path, mime=True))

    def create_student(self, offer_year, learning_unit_year, session_exam):
        student = StudentFactory()

        offer_enrollment = OfferEnrollmentFactory(offer_year=offer_year, student=student)
        learning_unit_enrollment = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment,
                                                                 learning_unit_year=learning_unit_year)
        enrollment = ExamEnrollmentFactory(learning_unit_enrollment=learning_unit_enrollment, session_exam=session_exam)

        return student, enrollment

    def create_user_person(self):
        user = self.create_user()
        self.add_group(user, 'program_managers', 'tutors')
        self.add_permission(user, 'can_access_academic_calendar', 'assessments.can_access_scoreencoding')
        person = PersonFactory(
            user=user,
            first_name=user.first_name,
            last_name=user.last_name,
            language='fr-be'
        )
        return user, person

    @classmethod
    def create_academic_year_calendar(self, year=None, start_date=None, days=None):
        if year is None:
            year = pendulum.today().year - 1

        academic_year = AcademicYearFactory(year=year)
        academic_calendar = AcademicCalendarExamSubmissionFactory.build(academic_year=academic_year)
        academic_calendar.save(functions=[])

        return academic_year, academic_calendar

    @classmethod
    def create_session_exam(self, learning_unit_year, session_exam_calendar, offer_year):
        return SessionExamFactory(
            learning_unit_year=learning_unit_year,
            number_session=session_exam_calendar.number_session,
            offer_year=offer_year
        )

    def create_exam_enrollments(self, offer_year, learning_unit_year, session_exam, number_of_students=10):
        return [
            self.create_student(offer_year, learning_unit_year, session_exam)[1]
            for counter in range(number_of_students)
        ]

    @classmethod
    def create_offers(cls, academic_year, acronyms, person=None):
        assert isinstance(acronyms, (list, tuple)) and len(acronyms) > 0
        offers = {
            acronym: OfferYearFactory(academic_year=academic_year, acronym=acronym)
            for acronym in acronyms
        }

        if person:
            for offer in offers.values():
                ProgramManagerFactory(offer_year=offer, person=person)

        return offers


import pypom

class Field:
    def __init__(self, *locator):
        self.locator = locator

class InputField(Field):
    def __set__(self, obj, value):
        element = obj.find_element(*self.locator)
        element.clear()
        if value is not None:
            element.send_keys(value)

    def __get__(self, obj, owner):
        element = obj.find_element(*self.locator)
        return element.get_attribute('value')

class SubmitField(Field):
    def __get__(self, obj, owner):
        return obj.find_element(*self.locator)

class ScoresEncodingPage(pypom.Page):
    acronym = InputField(By.ID, 'txt_acronym')
    search_button = SubmitField(By.ID, 'bt_submit_offer_search')
    via_paper = SubmitField(By.ID, 'lnk_via_paper')

    def search(self, acronym=None):
        self.acronym = acronym

        self.search_button.click()
