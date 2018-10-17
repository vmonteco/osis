##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2018 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
from django.utils.translation import ugettext_lazy as _

from base.models.external_learning_unit_year import ExternalLearningUnitYear
from base.models.learning_unit_year import LearningUnitYear
from base.models.proposal_learning_unit import ProposalLearningUnit
from osis_common.document import xls_build
from base.business.xls import get_name_or_username
from base.business.learning_unit_year_with_context import append_latest_entities, append_components, \
    get_learning_component_prefetch
from base.business.entity import build_entity_container_prefetch
from base.models.enums import entity_container_year_link_type as entity_types
from base.models.enums import learning_component_year_type
from base.business.learning_unit import get_organization_from_learning_unit_year
from base.business.learning_units.comparison import get_partims_as_str
from base.models.academic_year import current_academic_year
from openpyxl.utils import get_column_letter

# List of key that a user can modify
EMPTY_VALUE = ''
DATE_FORMAT = '%d-%m-%Y'
DATE_TIME_FORMAT = '%d-%m-%Y %H:%M'
DESC = "desc"
WORKSHEET_TITLE = 'learning_units_comparison'
XLS_FILENAME = 'learning_units_comparison'
XLS_DESCRIPTION = _("list_learning_units_comparison")

LEARNING_UNIT_TITLES = [
    str(_('code')),
    str(_('academic_year_small')),
    str(_('type')),
    str(_('active_title')),
    str(_('subtype')),
    str(_('Internship subtype')),
    str(_('credits')),
    str(_('language')),
    str(_('periodicity')),
    str(_('quadrimester')),
    str(_('session_title')),
    str(_('common_title')),
    str(_('title_proper_to_UE')),
    str(_('common_english_title')),
    str(_('english_title_proper_to_UE')),
    str(_('Req. Entities')),
    str(_('allocation_entity_small')),
    str(_('Add. requ. ent. 1')),
    str(_('Add. requ. ent. 2')),
    str(_('Profes. integration')),
    str(_('institution')),
    str(_('learning_location')),
    str(_('partims')),
    "PM {}".format(_('code')),
    "PM {}".format(_('volume_partial')),
    "PM {}".format(_('volume_remaining')),
    "PM {}".format(_('Vol. annual')),
    "PM {}".format(_('real_classes')),
    "PM {}".format(_('planned_classes')),
    "PM {}".format(_('vol_global')),
    "PM {}".format(_('Req. Entities')),
    "PM {}".format(_('Add. requ. ent. 1')),
    "PM {}".format(_('Add. requ. ent. 2')),
    "PP {}".format(_('code')),
    "PP {}".format(_('volume_partial')),
    "PP {}".format(_('volume_remaining')),
    "PP {}".format(_('Vol. annual')),
    "PM {}".format(_('real_classes')),
    "PM {}".format(_('planned_classes')),
    "PP {}".format(_('vol_global')),
    "PP {}".format(_('Req. Entities')),
    "PM {}".format(_('Add. requ. ent. 1')),
    "PM {}".format(_('Add. requ. ent. 2'))
]
ACRONYM_COL_NUMBER = 0
ACADEMIC_COL_NUMBER = 1
CELLS_MODIFIED_NO_BORDER = 'modifications'
CELLS_TOP_BORDER = 'border_not_modified'
DATA = 'data'


def create_xls_comparison(user, learning_unit_years, filters, academic_yr_comparison):
    working_sheets_data = []
    cells_modified_with_green_font = []
    cells_with_top_border = []

    if learning_unit_years:
        luys_for_2_years = _get_learning_unit_yrs_on_2_different_years(academic_yr_comparison, learning_unit_years)
        data = prepare_xls_content(luys_for_2_years)
        working_sheets_data = data.get('data')
        cells_modified_with_green_font = data.get(CELLS_MODIFIED_NO_BORDER)
        cells_with_top_border = data.get(CELLS_TOP_BORDER)

    parameters = {
        xls_build.DESCRIPTION: XLS_DESCRIPTION,
        xls_build.USER: get_name_or_username(user),
        xls_build.FILENAME: XLS_FILENAME,
        xls_build.HEADER_TITLES: LEARNING_UNIT_TITLES,
        xls_build.WS_TITLE: WORKSHEET_TITLE,
    }
    dict_styled_cells = {}
    if cells_modified_with_green_font:
        dict_styled_cells.update({xls_build.STYLE_MODIFIED: cells_modified_with_green_font})

    if cells_with_top_border:
        dict_styled_cells.update({xls_build.STYLE_BORDER_TOP: cells_with_top_border})
    if dict_styled_cells:
        parameters.update({xls_build.STYLED_CELLS: dict_styled_cells})
    return xls_build.generate_xls(xls_build.prepare_xls_parameters_list(working_sheets_data, parameters), filters)


def _get_learning_unit_yrs_on_2_different_years(academic_yr_comparison, learning_unit_years):
    learning_unit_years = \
        LearningUnitYear.objects.filter(
            learning_unit__in=(_get_learning_units(learning_unit_years)),
            academic_year__year__in=(
                learning_unit_years[0].academic_year.year,
                academic_yr_comparison)
        ).select_related(
            'academic_year',
            'learning_container_year',
            'learning_container_year__academic_year'
        ).prefetch_related(
            get_learning_component_prefetch()
        ).prefetch_related(
            build_entity_container_prefetch([
                entity_types.ALLOCATION_ENTITY,
                entity_types.REQUIREMENT_ENTITY,
                entity_types.ADDITIONAL_REQUIREMENT_ENTITY_1,
                entity_types.ADDITIONAL_REQUIREMENT_ENTITY_2
            ])
        ).order_by('learning_unit', 'academic_year__year')
    [append_latest_entities(learning_unit, False) for learning_unit in learning_unit_years]
    [append_components(learning_unit) for learning_unit in learning_unit_years]
    return learning_unit_years


def _get_learning_units(learning_unit_years):
    return list(set([l.learning_unit for l in learning_unit_years]))


def prepare_xls_content(learning_unit_yrs):
    data = []
    learning_unit = None
    first_data = None
    modified_cells_no_border = []
    top_border = []
    for line_index, l_u_yr in enumerate(learning_unit_yrs, start=1):

        if learning_unit is None:
            learning_unit = l_u_yr.learning_unit
            new_line = True
        else:
            if learning_unit == l_u_yr.learning_unit:
                new_line = False
            else:
                learning_unit = l_u_yr.learning_unit
                new_line = True
        luy_data = extract_xls_data_from_learning_unit(l_u_yr, new_line, first_data)
        if new_line:
            first_data = luy_data
            top_border.extend(get_border_columns(line_index+1))
        else:
            modified_cells_no_border.extend(_check_changes_other_than_code_and_year(first_data, luy_data, line_index+1))
            first_data = None
        data.append(luy_data)

    return {
        DATA: data,
        CELLS_TOP_BORDER: top_border or None,
        CELLS_MODIFIED_NO_BORDER: modified_cells_no_border or None,
    }


def extract_xls_data_from_learning_unit(learning_unit_yr, new_line, first_data):
    data = _get_data(learning_unit_yr, new_line, first_data)
    data.extend(_component_data(learning_unit_yr.components, learning_component_year_type.LECTURING))
    data.extend(_component_data(learning_unit_yr.components, learning_component_year_type.PRACTICAL_EXERCISES))
    return data


def _translate_status(value):
    if value:
        return _('active').title()
    else:
        return _('inactive').title()


def _component_data(components, learning_component_yr_type):
    if components:
        for component in components:
            if component.type == learning_component_yr_type:
                return _get_volumes(component, components)
    return [EMPTY_VALUE, EMPTY_VALUE, EMPTY_VALUE, EMPTY_VALUE, EMPTY_VALUE, EMPTY_VALUE, EMPTY_VALUE, EMPTY_VALUE,
            EMPTY_VALUE, EMPTY_VALUE]


def _get_data(learning_unit_yr, new_line, first_data):
    organization = get_organization_from_learning_unit_year(learning_unit_yr)
    return [
        _get_acronym(learning_unit_yr, new_line, first_data),
        learning_unit_yr.academic_year.name,
        xls_build.translate(learning_unit_yr.learning_container_year.container_type),
        _translate_status(learning_unit_yr.status),
        xls_build.translate(learning_unit_yr.subtype),
        _get_translation(learning_unit_yr.internship_subtype),
        learning_unit_yr.credits,
        learning_unit_yr.language.name if learning_unit_yr.language else EMPTY_VALUE,
        _get_translation(learning_unit_yr.periodicity),
        _get_translation(learning_unit_yr.quadrimester),
        _get_translation(learning_unit_yr.session),
        learning_unit_yr.learning_container_year.common_title,
        learning_unit_yr.specific_title,
        learning_unit_yr.learning_container_year.common_title_english,
        learning_unit_yr.specific_title_english,
        _get_entity_to_display(learning_unit_yr.entities.get(entity_types.REQUIREMENT_ENTITY)),
        _get_entity_to_display(learning_unit_yr.entities.get(entity_types.ALLOCATION_ENTITY)),
        _get_entity_to_display(learning_unit_yr.entities.get(entity_types.ADDITIONAL_REQUIREMENT_ENTITY_1)),
        _get_entity_to_display(learning_unit_yr.entities.get(entity_types.ADDITIONAL_REQUIREMENT_ENTITY_2)),
        xls_build.translate(learning_unit_yr.professional_integration),
        organization.name if organization else EMPTY_VALUE,
        learning_unit_yr.campus if learning_unit_yr.campus else EMPTY_VALUE,
        get_partims_as_str(learning_unit_yr.get_partims_related())
    ]


def _get_acronym(learning_unit_yr, new_line, first_data):
    acronym = EMPTY_VALUE
    if new_line:
        acronym = learning_unit_yr.acronym
    else:
        if learning_unit_yr.acronym != first_data[ACRONYM_COL_NUMBER]:
            acronym = learning_unit_yr.acronym
    return acronym


def _get_volumes(component, components):
    volumes = components[component]
    return [
        component.acronym if component.acronym else EMPTY_VALUE,
        volumes.get('VOLUME_Q1', EMPTY_VALUE),
        volumes.get('VOLUME_Q2', EMPTY_VALUE),
        volumes.get('VOLUME_TOTAL', EMPTY_VALUE),
        component.real_classes if component.real_classes else EMPTY_VALUE,
        component.planned_classes if component.planned_classes else EMPTY_VALUE,
        volumes.get('VOLUME_GLOBAL', '0'),
        volumes.get('VOLUME_REQUIREMENT_ENTITY', EMPTY_VALUE),
        volumes.get('VOLUME_ADDITIONAL_REQUIREMENT_ENTITY_1', EMPTY_VALUE),
        volumes.get('VOLUME_ADDITIONAL_REQUIREMENT_ENTITY_2', EMPTY_VALUE)
    ]


def _get_translation(value):
    return str(_(value)) if value else EMPTY_VALUE


def _get_entity_to_display(entity):
    return entity.acronym if entity else EMPTY_VALUE


def get_academic_year_of_reference(objects):
    """ TODO : Has to be improved because it's not optimum if the xls list is created from a search with a
    criteria : academic_year = 'ALL' """
    if objects:
        return _get_academic_year(objects[0])
    return current_academic_year()


def _get_academic_year(obj):
    if isinstance(obj, LearningUnitYear):
        return obj.academic_year
    if isinstance(obj, (ProposalLearningUnit, ExternalLearningUnitYear)):
        return obj.learning_unit_year.academic_year


def _check_changes_other_than_code_and_year(first_data, second_data, line_index):
    modifications = []
    for col_index, obj in enumerate(first_data):
        if col_index == ACRONYM_COL_NUMBER and second_data[ACRONYM_COL_NUMBER] != EMPTY_VALUE:
            modifications.append('{}{}'.format(get_column_letter(col_index+1), line_index))
        else:
            if obj != second_data[col_index] and col_index != ACADEMIC_COL_NUMBER:
                modifications.append('{}{}'.format(get_column_letter(col_index+1), line_index))

    return modifications


def get_border_columns(line):
    style = []
    for col_index, obj in enumerate(LEARNING_UNIT_TITLES, start=1):
        style.append('{}{}'.format(get_column_letter(col_index), line))
    return style
