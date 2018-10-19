##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2018 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################

from django.utils.translation import ugettext_lazy as _

from base import models as mdl_base
from base.business.learning_unit import LEARNING_UNIT_TITLES_PART2,  XLS_DESCRIPTION, XLS_FILENAME, \
    WORKSHEET_TITLE, get_same_container_year_components, get_entity_acronym
from base.business.xls import get_name_or_username
from osis_common.document import xls_build
from attribution.business import attribution_charge_new
from base.models.enums import learning_component_year_type
from openpyxl.styles import Alignment, Style, PatternFill, Color, Font
from base.models.enums.proposal_type import ProposalType
from openpyxl.utils import get_column_letter
from collections import defaultdict
# List of key that a user can modify
VOLUMES_INITIALIZED = {'VOLUME_TOTAL': 0, 'PLANNED_CLASSES': 0, 'VOLUME_Q1': 0, 'VOLUME_Q2': 0}

TRANSFORMATION_AND_MODIFICATION_COLOR = Color('808000')
TRANSFORMATION_COLOR = Color('ff6600')
SUPPRESSION_COLOR = Color('ff0000')
MODIFICATION_COLOR = Color('0000ff')
CREATION_COLOR = Color('008000')
DEFAULT_LEGEND_STYLES = {
    Style(fill=PatternFill(patternType='solid', fgColor=CREATION_COLOR)): ['A2'],
    Style(fill=PatternFill(patternType='solid', fgColor=MODIFICATION_COLOR)): ['A3'],
    Style(fill=PatternFill(patternType='solid', fgColor=SUPPRESSION_COLOR)): ['A4'],
    Style(fill=PatternFill(patternType='solid', fgColor=TRANSFORMATION_COLOR)): ['A5'],
    Style(fill=PatternFill(patternType='solid', fgColor=TRANSFORMATION_AND_MODIFICATION_COLOR)): ['A6'],
}
SPACES = '  '
HEADER_TEACHERS = _('List of teachers')
HEADER_PROGRAMS = _('Programs')
PROPOSAL_LINE_STYLES = {
    ProposalType.CREATION.name: Style(font=Font(color=CREATION_COLOR),),
    ProposalType.MODIFICATION.name: Style(font=Font(color=MODIFICATION_COLOR),),
    ProposalType.SUPPRESSION.name: Style(font=Font(color=SUPPRESSION_COLOR),),
    ProposalType.TRANSFORMATION.name: Style(font=Font(color=TRANSFORMATION_COLOR),),
    ProposalType.TRANSFORMATION_AND_MODIFICATION.name: Style(font=Font(color=TRANSFORMATION_AND_MODIFICATION_COLOR),),
}
WRAP_TEXT_STYLE = Style(alignment=Alignment(wrapText=True, vertical="top"), )
WITH_ATTRIBUTIONS = 'with_attributions'
WITH_GRP = 'with_grp'

LEARNING_UNIT_TITLES_PART1 = [
    str(_('code')),
    str(_('academic_year_small')),
    str(_('title')),
    str(_('type')),
    str(_('subtype')),
    "{} ({})".format(_('requirement_entity_small'), _('fac. level')),
    str(_('proposal_type')),
    str(_('proposal_status')),
    str(_('credits')),
    str(_('allocation_entity_small')),
    str(_('title_in_english')),
]


def prepare_xls_content(learning_units,
                        with_grp=False,
                        with_attributions=False):
    return [
        extract_xls_data_from_learning_unit(lu,
                                            with_grp,
                                            with_attributions)
        for lu in learning_units
        ]


def extract_xls_data_from_learning_unit(learning_unit_yr, with_grp, with_attributions):

    lu_data_part1 = _get_data_part1(learning_unit_yr)
    lu_data_part2 = _get_data_part2(learning_unit_yr, with_attributions)

    if with_grp:
        lu_data_part2.append(_add_training_data(learning_unit_yr))
    lu_data_part1.extend(lu_data_part2)
    return lu_data_part1


def create_xls_with_parameters(user, learning_units, filters, extra_configuration):
    with_grp = extra_configuration.get(WITH_GRP)
    with_attributions = extra_configuration.get(WITH_ATTRIBUTIONS)
    titles_part1 = LEARNING_UNIT_TITLES_PART1.copy()
    titles_part2 = LEARNING_UNIT_TITLES_PART2.copy()

    if with_grp:
        titles_part2.append(str(HEADER_PROGRAMS))

    if with_attributions:
        titles_part1.append(str(HEADER_TEACHERS))
        for learning_unit_yr in learning_units:
            learning_unit_yr.attribution_charge_news = attribution_charge_new \
                .find_attribution_charge_new_by_learning_unit_year_as_dict(learning_unit_year=learning_unit_yr)

    working_sheets_data = prepare_xls_content(learning_units, with_grp, with_attributions)

    titles_part1.extend(titles_part2)

    ws_data = xls_build.prepare_xls_parameters_list(working_sheets_data,
                                                    _get_parameters_configurable_list(learning_units,
                                                                                      titles_part1,
                                                                                      user))
    ws_data.update({xls_build.WORKSHEETS_DATA: [ws_data.get(xls_build.WORKSHEETS_DATA)[0], _prepare_legend_ws_data()]})
    return xls_build.generate_xls(ws_data, filters)


def _get_parameters_configurable_list(learning_units, titles, user):
    parameters = {
        xls_build.DESCRIPTION: XLS_DESCRIPTION,
        xls_build.USER: get_name_or_username(user),
        xls_build.FILENAME: XLS_FILENAME,
        xls_build.HEADER_TITLES: titles,
        xls_build.WS_TITLE: WORKSHEET_TITLE,
        xls_build.STYLED_CELLS: {
            WRAP_TEXT_STYLE: _get_wrapped_cells(
                learning_units,
                _get_col_letter(titles, HEADER_PROGRAMS),
                _get_col_letter(titles, HEADER_TEACHERS)
            )
        },
        xls_build.COLORED_ROWS: _get_colored_rows(learning_units),
    }
    return parameters


def _get_absolute_credits(learning_unit_yr):
    group_elements_years = mdl_base.group_element_year.search(child_leaf=learning_unit_yr) \
        .select_related("child_leaf", "parent__education_group_type").order_by('parent__partial_acronym')
    if group_elements_years:
        return group_elements_years.first().child_leaf.credits \
            if group_elements_years.first().child_leaf.credits else ''
    return ''


def _get_volumes(learning_unit_yr):
    volumes = {
        learning_component_year_type.LECTURING: VOLUMES_INITIALIZED,
        learning_component_year_type.PRACTICAL_EXERCISES: VOLUMES_INITIALIZED
    }
    data_components = get_same_container_year_components(learning_unit_yr, True)
    for component in data_components.get('components'):
        if component.get('learning_component_year').type in (learning_component_year_type.LECTURING,
                                                             learning_component_year_type.PRACTICAL_EXERCISES):
            volumes = _update_volumes_data(component, volumes)

    return volumes


def _update_volumes_data(component, volumes_param):
    volumes = volumes_param.copy()
    vol_to_update = volumes.get(component.get('learning_component_year').type).copy()
    key_of_value_to_update = ['VOLUME_TOTAL', 'VOLUME_Q1', 'VOLUME_Q2', 'PLANNED_CLASSES']
    for key in key_of_value_to_update:
        if component.get('volumes').get(key):
            vol_to_update[key] = vol_to_update.get(key) + component.get('volumes').get(key)

    volumes[component.get('learning_component_year').type] = vol_to_update
    return volumes


def _get_significant_volume(volume):
    if volume and volume > 0:
        return volume
    return ''


def _prepare_legend_ws_data():
    return {
        xls_build.HEADER_TITLES_KEY: [str(_('Legend'))],
        xls_build.CONTENT_KEY: [
            [SPACES, _('proposal_creation')],
            [SPACES, _('Proposal for modification')],
            [SPACES, _('Suppression proposal')],
            [SPACES, _('Transformation proposal')],
            [SPACES, _('Transformation/modification proposal')],
        ],
        xls_build.WORKSHEET_TITLE_KEY: _('Legend'),
        xls_build.STYLED_CELLS:
            DEFAULT_LEGEND_STYLES
    }


def _get_wrapped_cells(learning_units, teachers_col_letter, programs_col_letter):
    dict_wrapped_styled_cells = []

    for idx, luy in enumerate(learning_units, start=2):
        if teachers_col_letter:
            dict_wrapped_styled_cells.append("{}{}".format(teachers_col_letter, idx))
        if programs_col_letter:
            dict_wrapped_styled_cells.append("{}{}".format(programs_col_letter, idx))

    return dict_wrapped_styled_cells


def _get_colored_rows(learning_units):
    colored_cells = defaultdict(list)
    for idx, luy in enumerate(learning_units, start=1):
        proposal = mdl_base.proposal_learning_unit.find_by_learning_unit_year(luy)
        if proposal:
            colored_cells[PROPOSAL_LINE_STYLES.get(proposal.type)].append(idx)
    return colored_cells


def _get_attribution_line(an_attribution):
    return "{} - {} : {} - {} : {} - {} : {} - {} : {} - {} : {} - {} : {} ".format(
        an_attribution.get('person'),
        _('function'),
        _(an_attribution.get('function')) if an_attribution.get('function') else '',
        _('substitute'),
        an_attribution.get('substitute') if an_attribution.get('substitute') else '',
        _('Beg. of attribution'),
        an_attribution.get('start_year'),
        _('Attribution duration'),
        an_attribution.get('duration'),
        _('Attrib. vol1'),
        an_attribution.get('LECTURING'),
        _('Attrib. vol2'),
        an_attribution.get('PRACTICAL_EXERCISES'),
    )


def _get_col_letter(titles, title_search):
    for idx, title in enumerate(titles, start=1):
        if title == title_search:
            return get_column_letter(idx)
    return None


def _get_trainings_by_educ_group_year(learning_unit_yr):
    groups = []
    learning_unit_yr.group_elements_years = mdl_base.group_element_year.search(child_leaf=learning_unit_yr) \
        .select_related("child_leaf", "parent__education_group_type") \
        .order_by('parent__partial_acronym')
    groups.extend(learning_unit_yr.group_elements_years)
    education_groups_years = [group_element_year.parent for group_element_year in groups]
    return mdl_base.group_element_year \
        .find_learning_unit_formations(education_groups_years, parents_as_instances=True)


def _add_training_data(learning_unit_yr):
    formations_by_educ_group_year = _get_trainings_by_educ_group_year(learning_unit_yr)
    return " \n".join(["{}".format(_concatenate_training_data(formations_by_educ_group_year, group_element_year)) for
                       group_element_year in learning_unit_yr.group_elements_years])


def _concatenate_training_data(formations_by_educ_group_year, group_element_year):
    training_string = "{} {} {}".format(
        group_element_year.parent.partial_acronym if group_element_year.parent.partial_acronym else '',
        "({})".format(
            '{0:.2f}'.format(group_element_year.child_leaf.credits) if group_element_year.child_leaf.credits else '-'),
        " - ".join(
            ["{} - {}".format(training.acronym, training.title) for training in
             formations_by_educ_group_year.get(group_element_year.parent_id)])
    )
    return training_string


def _get_data_part2(learning_unit_yr, with_attributions):
    volumes = _get_volumes(learning_unit_yr)
    lu_data_part2 = []
    if with_attributions:
        lu_data_part2.append(
            " \n".join([_get_attribution_line(value) for value in learning_unit_yr.attribution_charge_news.values()]))
    volume_lecturing = volumes.get(learning_component_year_type.LECTURING)
    volumes_practical = volumes.get(learning_component_year_type.PRACTICAL_EXERCISES)
    lu_data_part2.extend([
        xls_build.translate(learning_unit_yr.periodicity),
        xls_build.translate(learning_unit_yr.status),
        _get_significant_volume(volume_lecturing.get('VOLUME_TOTAL')),
        _get_significant_volume(volume_lecturing.get('VOLUME_Q1')),
        _get_significant_volume(volume_lecturing.get('VOLUME_Q2')),
        _get_significant_volume(volume_lecturing.get('PLANNED_CLASSES')),
        _get_significant_volume(volumes_practical.get('VOLUME_TOTAL')),
        _get_significant_volume(volumes_practical.get('VOLUME_Q1')),
        _get_significant_volume(volumes_practical.get('VOLUME_Q2')),
        _get_significant_volume(volumes_practical.get('PLANNED_CLASSES')),
        xls_build.translate(learning_unit_yr.quadrimester),
        xls_build.translate(learning_unit_yr.session),
        learning_unit_yr.language if learning_unit_yr.language else "",
        _get_absolute_credits(learning_unit_yr),
    ])
    return lu_data_part2


def _get_data_part1(learning_unit_yr):
    proposal = mdl_base.proposal_learning_unit.find_by_learning_unit_year(learning_unit_yr)
    lu_data_part1 = [
        learning_unit_yr.acronym,
        learning_unit_yr.academic_year.name,
        learning_unit_yr.complete_title,
        xls_build.translate(learning_unit_yr.learning_container_year.container_type)
        # FIXME Condition to remove when the LearningUnitYear.learning_continer_year_id will be null=false
        if learning_unit_yr.learning_container_year else "",
        xls_build.translate(learning_unit_yr.subtype),
        _get_entity_faculty_acronym(learning_unit_yr.entities.get('REQUIREMENT_ENTITY'),
                                    learning_unit_yr.academic_year),
        xls_build.translate(proposal.type) if proposal else '',
        xls_build.translate(proposal.state) if proposal else '',
        learning_unit_yr.credits,
        get_entity_acronym(learning_unit_yr.entities.get('ALLOCATION_ENTITY')),
        learning_unit_yr.complete_title_english,
    ]
    return lu_data_part1


def _get_entity_faculty_acronym(an_entity, academic_yr):
    if an_entity:
        faculty_entity = an_entity.find_faculty_version(academic_yr)
        return faculty_entity.acronym if faculty_entity else None
    return None
