##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2018 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
from collections import OrderedDict
from operator import itemgetter

from django.conf import settings
from django.utils.translation import ugettext_lazy as _

from attribution.models import attribution
from attribution.models.attribution import find_all_tutors_by_learning_unit_year
from base import models as mdl_base
from base.business.entity import get_entity_calendar
from base.business.learning_unit_year_with_context import volume_learning_component_year
from base.business.learning_units.comparison import get_entity_by_type
from base.business.xls import get_name_or_username
from base.models import entity_container_year, academic_calendar, proposal_learning_unit
from base.models import learning_achievement
from base.models.entity_component_year import EntityComponentYear
from base.models.enums import academic_calendar_type
from base.models.enums.entity_container_year_link_type import REQUIREMENT_ENTITIES
from cms import models as mdl_cms
from cms.enums import entity_name
from cms.enums.entity_name import LEARNING_UNIT_YEAR
from cms.models import translated_text
from osis_common.document import xls_build
from osis_common.utils.datetime import convert_date_to_datetime
from base.models.enums import entity_container_year_link_type
from attribution.business import attribution_charge_new
from base.models.enums import learning_component_year_type
from openpyxl.styles import Alignment, Style, PatternFill, Color, Font
from base.models.enums import proposal_type, proposal_state

# List of key that a user can modify
TRANSFORMATION_AND_MODIFICATION = Color('808000')
TRANSFORMATION_COLOR = Color('ff6600')
SUPPRESSION_COLOR = Color('ff0000')
MODIFICATION_COLOR = Color('0000ff')
CREATION_COLOR = Color('008000')
DASH = '-'
EMPTY = ''
WORKSHEET_TITLE = 'learning_units'
XLS_FILENAME = 'learning_units_filename'
XLS_DESCRIPTION = "List_activities"
LEARNING_UNIT_TITLES_PART1 = [
    str(_('code')),
    str(_('academic_year_small')),
    str(_('title')),
    str(_('type')),
    str(_('subtype')),
    str(_('requirement_entity_small')),
    str(_('proposal_type')),
    str(_('proposal_status')),
    str(_('credits')),
    str(_('allocation_entity_small')),
    str(_('title_in_english')),
]

LEARNING_UNIT_TITLES_PART2 = [
    str(_('periodicity')),
    str(_('active_title')),
    "{} 1 - {}".format(str(_('Hourly vol.')), str(_('ANNUAL'))),
    "{} 1 - {}".format(str(_('Hourly vol.')), str(_('1st quadri'))),
    "{} 1 - {}".format(str(_('Hourly vol.')), str(_('2nd quadri'))),
    "{} 1".format(str(_('PLANNED_CLASSES'))),
    "{} 2 - {}".format(str(_('Hourly vol.')), str(_('ANNUAL'))),
    "{} 2 - {}".format(str(_('Hourly vol.')), str(_('1st quadri'))),
    "{} 2 - {}".format(str(_('Hourly vol.')), str(_('2nd quadri'))),
    "{} 2".format(str(_('PLANNED_CLASSES'))),
    str(_('quadrimester')),
    str(_('session_title')),
    str(_('language')),
    str(_('Absolute credits')),
]
CMS_LABEL_SPECIFICATIONS = ['themes_discussed', 'prerequisite']

CMS_LABEL_PEDAGOGY_FR_AND_EN = ['resume', 'teaching_methods', 'evaluation_methods', 'other_informations',
                                'online_resources']
CMS_LABEL_PEDAGOGY_FR_ONLY = ['bibliography', 'mobility']
CMS_LABEL_PEDAGOGY = CMS_LABEL_PEDAGOGY_FR_AND_EN + CMS_LABEL_PEDAGOGY_FR_ONLY

CMS_LABEL_SUMMARY = ['resume']
CARRIAGE_RETURN = " \n"

PROPOSAL_LINE_STYLES = {
    proposal_type.ProposalType.CREATION.name: Style(font=Font(color=CREATION_COLOR),),
    proposal_type.ProposalType.MODIFICATION.name: Style(font=Font(color=MODIFICATION_COLOR),),
    proposal_type.ProposalType.SUPPRESSION.name: Style(font=Font(color=SUPPRESSION_COLOR),),
    proposal_type.ProposalType.TRANSFORMATION.name: Style(font=Font(color=TRANSFORMATION_COLOR),),
    proposal_type.ProposalType.TRANSFORMATION_AND_MODIFICATION.name:
        Style(font=Font(color=TRANSFORMATION_AND_MODIFICATION),),
}
COLORED = 'COLORED_ROW'
WRAP = 'WRAP'
WRAP_TEXT_STYLE = Style(alignment=Alignment(wrapText=True, vertical="top"), )
WITH_ATTRIBUTIONS = 'with_attributions'
WITH_GRP = 'with_grp'


def get_same_container_year_components(learning_unit_year, with_classes=False):
    learning_container_year = learning_unit_year.learning_container_year
    components = []
    learning_components_year = mdl_base.learning_component_year.find_by_learning_container_year(learning_container_year,
                                                                                                with_classes)
    additionnal_entities = {}

    for indx, learning_component_year in enumerate(learning_components_year):
        if learning_component_year.classes:
            for learning_class_year in learning_component_year.classes:
                learning_class_year.used_by_learning_units_year = _learning_unit_usage_by_class(learning_class_year)
                learning_class_year.is_used_by_full_learning_unit_year = _is_used_by_full_learning_unit_year(
                    learning_class_year)

        used_by_learning_unit = mdl_base.learning_unit_component.search(learning_component_year, learning_unit_year)

        entity_components_yr = EntityComponentYear.objects.filter(learning_component_year=learning_component_year)
        if indx == 0:
            additionnal_entities = _get_entities(entity_components_yr)

        components.append({'learning_component_year': learning_component_year,
                           'volumes': volume_learning_component_year(learning_component_year, entity_components_yr),
                           'learning_unit_usage': _learning_unit_usage(learning_component_year),
                           'used_by_learning_unit': used_by_learning_unit
                           })

    components = sorted(components, key=itemgetter('learning_unit_usage'))
    return _compose_components_dict(components, additionnal_entities)


def get_organization_from_learning_unit_year(learning_unit_year):
    if learning_unit_year.campus:
        return learning_unit_year.campus.organization


def get_all_attributions(learning_unit_year):
    attributions = {}
    if learning_unit_year.learning_container_year:
        all_attributions = entity_container_year.find_last_entity_version_grouped_by_linktypes(
            learning_unit_year.learning_container_year)

        attributions['requirement_entity'] = all_attributions.get(entity_container_year_link_type.REQUIREMENT_ENTITY)
        attributions['allocation_entity'] = all_attributions.get(entity_container_year_link_type.ALLOCATION_ENTITY)
        attributions['additional_requirement_entity_1'] = \
            all_attributions.get(entity_container_year_link_type.ADDITIONAL_REQUIREMENT_ENTITY_1)
        attributions['additional_requirement_entity_2'] = \
            all_attributions.get(entity_container_year_link_type.ADDITIONAL_REQUIREMENT_ENTITY_2)
    return attributions


def get_cms_label_data(cms_label, user_language):
    cms_label_data = OrderedDict()
    translated_labels = mdl_cms.translated_text_label.search(text_entity=entity_name.LEARNING_UNIT_YEAR,
                                                             labels=cms_label,
                                                             language=user_language)

    for label in cms_label:
        translated_text = next((trans.label for trans in translated_labels if trans.text_label.label == label), None)
        cms_label_data[label] = translated_text

    return cms_label_data


def _learning_unit_usage(a_learning_component_year):
    components = mdl_base.learning_unit_component.find_by_learning_component_year(a_learning_component_year)
    return ", ".join(["{} ({})".format(c.learning_unit_year.acronym, _(c.learning_unit_year.quadrimester or '?'))
                      for c in components])


def _learning_unit_usage_by_class(a_learning_class_year):
    queryset = mdl_base.learning_unit_component_class.find_by_learning_class_year(a_learning_class_year) \
        .order_by('learning_unit_component__learning_unit_year__acronym') \
        .values_list('learning_unit_component__learning_unit_year__acronym', flat=True)
    return ", ".join(list(queryset))


def get_components_identification(learning_unit_yr):
    a_learning_container_yr = learning_unit_yr.learning_container_year
    components = []
    additionnal_entities = {}

    if a_learning_container_yr:
        learning_component_year_list = mdl_base.learning_component_year.find_by_learning_container_year(
            a_learning_container_yr)

        for learning_component_year in learning_component_year_list:
            if mdl_base.learning_unit_component.search(learning_component_year, learning_unit_yr).exists():
                entity_components_yr = EntityComponentYear.objects.filter(
                    learning_component_year=learning_component_year)
                if not additionnal_entities:
                    additionnal_entities = _get_entities(entity_components_yr)

                components.append({'learning_component_year': learning_component_year,
                                   'entity_component_yr': entity_components_yr.first(),
                                   'volumes': volume_learning_component_year(learning_component_year,
                                                                             entity_components_yr)})

    return _compose_components_dict(components, additionnal_entities)


def _is_used_by_full_learning_unit_year(a_learning_class_year):
    for l in mdl_base.learning_unit_component_class.find_by_learning_class_year(a_learning_class_year):
        if l.learning_unit_component.learning_unit_year.subdivision is None:
            return True

    return False


def prepare_xls_content(found_learning_units,
                        with_grp=False,
                        with_attributions=False,
                        formations_by_educ_group_year=None):
    return [
        extract_xls_data_from_learning_unit(lu,
                                            with_grp,
                                            with_attributions,
                                            formations_by_educ_group_year)
        for lu in found_learning_units
        ]


def extract_xls_data_from_learning_unit(learning_unit_yr, with_grp=False, with_attributions=False,
                                        formations_by_educ_group_year=None):

    volumes = _get_volumes(learning_unit_yr)
    proposal = mdl_base.proposal_learning_unit.find_by_learning_unit_year(learning_unit_yr)

    lu_data_part1 = [
        learning_unit_yr.acronym,
        learning_unit_yr.academic_year.name,
        learning_unit_yr.complete_title.lstrip(),
        xls_build.translate(learning_unit_yr.learning_container_year.container_type)
        # FIXME Condition to remove when the LearningUnitYear.learning_continer_year_id will be null=false
        if learning_unit_yr.learning_container_year else "",
        xls_build.translate(learning_unit_yr.subtype),
        get_entity_acronym(learning_unit_yr.entities.get('REQUIREMENT_ENTITY')),
        xls_build.translate(proposal.type) if proposal else '',
        xls_build.translate(proposal.state) if proposal else '',
        learning_unit_yr.credits,
        get_entity_acronym(learning_unit_yr.entities.get('ALLOCATION_ENTITY')),
        learning_unit_yr.complete_title_english.lstrip(),
        ]
    lu_data_part2 = []
    if with_attributions:
        lu_data_part2.append(_append_attributions(learning_unit_yr))

    lu_data_part2.extend([
        xls_build.translate(learning_unit_yr.periodicity),
        xls_build.translate(learning_unit_yr.status),
        _get_significant_volume(volumes.get(learning_component_year_type.LECTURING).get('VOLUME_TOTAL')),
        _get_significant_volume(volumes.get(learning_component_year_type.LECTURING).get('VOLUME_Q1')),
        _get_significant_volume(volumes.get(learning_component_year_type.LECTURING).get('VOLUME_Q2')),
        _get_significant_volume(volumes.get(learning_component_year_type.LECTURING).get('PLANNED_CLASSES')),
        _get_significant_volume(volumes.get(learning_component_year_type.PRACTICAL_EXERCISES).get('VOLUME_TOTAL')),
        _get_significant_volume(volumes.get(learning_component_year_type.PRACTICAL_EXERCISES).get('VOLUME_Q1')),
        _get_significant_volume(volumes.get(learning_component_year_type.PRACTICAL_EXERCISES).get('VOLUME_Q2')),
        _get_significant_volume(volumes.get(learning_component_year_type.PRACTICAL_EXERCISES).get('PLANNED_CLASSES')),
        xls_build.translate(learning_unit_yr.quadrimester),
        xls_build.translate(learning_unit_yr.session),
        learning_unit_yr.language if learning_unit_yr.language else "",
        _get_absolute_credits(learning_unit_yr),
    ])

    if with_grp:
        ch = EMPTY
        for cpt, group_element_year in enumerate(learning_unit_yr.group_elements_years):
            if cpt > 0:
                ch += CARRIAGE_RETURN
            ch = "{} {} {} {}".format(
                ch,
                group_element_year.parent.partial_acronym if group_element_year.parent.partial_acronym else EMPTY,
                "({})".format(
                    str(group_element_year.child_leaf.credits) if group_element_year.child_leaf.credits else DASH),
                _get_trainings(group_element_year, formations_by_educ_group_year)
            )
        lu_data_part2.append(ch)
    lu_data_part1.extend(lu_data_part2)
    return lu_data_part1


def _append_attributions(learning_unit_yr):
    attributions_line = ''
    cpt = 0
    for key, value in learning_unit_yr.attribution_charge_news.items():
        if cpt > 0:
            attributions_line += CARRIAGE_RETURN
        attributions_line = _get_attribution_line(attributions_line, value)
        cpt += 1
    return attributions_line


def get_entity_acronym(an_entity):
    return an_entity.acronym if an_entity else None


def create_xls(user, found_learning_units, filters):
    titles = LEARNING_UNIT_TITLES_PART1.copy()
    titles.extend(LEARNING_UNIT_TITLES_PART2.copy())
    working_sheets_data = prepare_xls_content(found_learning_units, False, False, None)
    parameters = {xls_build.DESCRIPTION: XLS_DESCRIPTION,
                  xls_build.USER: get_name_or_username(user),
                  xls_build.FILENAME: XLS_FILENAME,
                  xls_build.HEADER_TITLES: titles,
                  xls_build.WS_TITLE: WORKSHEET_TITLE}

    return xls_build.generate_xls(xls_build.prepare_xls_parameters_list(working_sheets_data, parameters), filters)


def is_summary_submission_opened():
    current_academic_year = mdl_base.academic_year.current_academic_year()
    return mdl_base.academic_calendar.\
        is_academic_calendar_opened_for_specific_academic_year(current_academic_year,
                                                               academic_calendar_type.SUMMARY_COURSE_SUBMISSION)


def find_language_in_settings(language_code):
    return next((lang for lang in settings.LANGUAGES if lang[0] == language_code), None)


def _compose_components_dict(components, additional_entities):
    data_components = {'components': components}
    data_components.update(additional_entities)
    return data_components


def _get_entities(entity_components_yr):
    return {e.entity_container_year.type: e.entity_container_year.entity.most_recent_acronym
            for e in entity_components_yr
            if e.entity_container_year.type in REQUIREMENT_ENTITIES}


def _get_summary_status(a_calendar, cms_list, lu):
    for educational_information in cms_list:
        if educational_information.reference == lu.id \
                and _changed_in_period(a_calendar.start_date, educational_information.changed):
            return True
    return False


def _get_calendar(academic_yr, an_entity_version):
    a_calendar = get_entity_calendar(an_entity_version, academic_yr)
    if a_calendar is None:
        a_calendar = academic_calendar.get_by_reference_and_academic_year(
            academic_calendar_type.SUMMARY_COURSE_SUBMISSION,
            academic_yr
        )
    return a_calendar


def _get_summary_detail(a_calendar, cms_list, entity_learning_unit_yr_list_param):
    entity_learning_unit_yr_list = entity_learning_unit_yr_list_param
    for lu in entity_learning_unit_yr_list:
        lu.summary_responsibles = attribution.search(summary_responsible=True,
                                                     learning_unit_year=lu)
        lu.summary_status = _get_summary_status(a_calendar, cms_list, lu)
    return entity_learning_unit_yr_list


def _changed_in_period(start_date, changed_date):
    return convert_date_to_datetime(start_date) <= changed_date


def get_learning_units_and_summary_status(learning_unit_years):
    learning_units_found = []
    cms_list = translated_text.find_with_changed(LEARNING_UNIT_YEAR, CMS_LABEL_PEDAGOGY)
    for learning_unit_yr in learning_unit_years:
        learning_units_found.extend(_get_learning_unit_by_luy_entity(cms_list, learning_unit_yr))
    return learning_units_found


def _get_learning_unit_by_luy_entity(cms_list, learning_unit_yr):
    requirement_entity = learning_unit_yr.entities.get('REQUIREMENT_ENTITY', None)
    if requirement_entity:
        a_calendar = _get_calendar(learning_unit_yr.academic_year.past(), requirement_entity)
        if a_calendar:
            return _get_summary_detail(a_calendar, cms_list, [learning_unit_yr])
    return []


def get_achievements_group_by_language(learning_unit_year):
    achievement_grouped = {}
    all_achievements = learning_achievement.find_by_learning_unit_year(learning_unit_year)
    for achievement in all_achievements:
        key = 'achievements_{}'.format(achievement.language.code)
        achievement_grouped.setdefault(key, []).append(achievement)
    return achievement_grouped


def get_no_summary_responsible_teachers(learning_unit_yr, summary_responsibles):
    tutors = find_all_tutors_by_learning_unit_year(learning_unit_yr, "-summary_responsible")
    return [tutor[0] for tutor in tutors if tutor[0] not in summary_responsibles]


def get_learning_unit_comparison_context(learning_unit_year):
    context = dict({'learning_unit_year': learning_unit_year})
    context['campus'] = learning_unit_year.campus
    context['organization'] = get_organization_from_learning_unit_year(learning_unit_year)
    context['experimental_phase'] = True
    components = get_components_identification(learning_unit_year)
    context['components'] = components.get('components')
    context['REQUIREMENT_ENTITY'] = get_entity_by_type(learning_unit_year,
                                                       entity_container_year_link_type.REQUIREMENT_ENTITY)
    context['ALLOCATION_ENTITY'] = get_entity_by_type(learning_unit_year,
                                                      entity_container_year_link_type.ALLOCATION_ENTITY)
    context['ADDITIONAL_REQUIREMENT_ENTITY_1'] = \
        get_entity_by_type(learning_unit_year, entity_container_year_link_type.ADDITIONAL_REQUIREMENT_ENTITY_1)
    context['ADDITIONAL_REQUIREMENT_ENTITY_2'] = \
        get_entity_by_type(learning_unit_year, entity_container_year_link_type.ADDITIONAL_REQUIREMENT_ENTITY_2)
    context['learning_container_year_partims'] = learning_unit_year.get_partims_related()
    return context


def create_xls_with_parameters(user, found_learning_units, filters, extra_configuration):
    with_grp = extra_configuration.get(WITH_GRP)
    with_attributions = extra_configuration.get(WITH_ATTRIBUTIONS)
    titles_part1 = LEARNING_UNIT_TITLES_PART1.copy()
    titles_part2 = LEARNING_UNIT_TITLES_PART2.copy()

    formations_by_educ_group_year = []
    if with_grp:
        titles_part2.append(str(_('Programs')))
        formations_by_educ_group_year = _get_formations_by_educ_group_year(found_learning_units)
    if with_attributions:
        titles_part1.append(str(_('List of teachers')))
        for learning_unit_yr in found_learning_units:
            learning_unit_yr.attribution_charge_news = attribution_charge_new \
                .find_attribution_charge_new_by_learning_unit_year(learning_unit_year=learning_unit_yr)

    working_sheets_data = prepare_xls_content(
        found_learning_units,
        with_grp,
        with_attributions,
        formations_by_educ_group_year
    )

    titles_part1.extend(titles_part2)

    ws_data = xls_build.prepare_xls_parameters_list(working_sheets_data,
                                                    _get_parameters_configurable_list(found_learning_units,
                                                                                      titles_part1,
                                                                                      user))
    ws_data.update({xls_build.WORKSHEETS_DATA: [ws_data.get(xls_build.WORKSHEETS_DATA)[0], _prepare_legend_ws_data()]})
    return xls_build.generate_xls(ws_data, filters)


def _get_parameters_configurable_list(found_learning_units, titles_part1, user):
    dict_styles = _get_format(found_learning_units)
    parameters = {
        xls_build.DESCRIPTION: XLS_DESCRIPTION,
        xls_build.USER: get_name_or_username(user),
        xls_build.FILENAME: XLS_FILENAME,
        xls_build.HEADER_TITLES: titles_part1,
        xls_build.WS_TITLE: WORKSHEET_TITLE,
        xls_build.STYLED_CELLS: dict_styles.get(WRAP),
        xls_build.COLORED_ROWS: dict_styles.get(COLORED)
    }
    return parameters


def _get_trainings(group_element_year, formations_by_educ_group_year):
    ch = ''
    for cpt, training in enumerate(formations_by_educ_group_year.get(group_element_year.parent_id)):
        if cpt > 0:
            ch += " - "
        ch = "{} {} - {} ".format(ch, training.acronym, training.title)
    return ch


def _get_absolute_credits(learning_unit_yr):
    group_elements_years = mdl_base.group_element_year.search(child_leaf=learning_unit_yr) \
        .select_related("parent", "child_leaf", "parent__education_group_type").order_by('parent__partial_acronym')
    if group_elements_years:
        return group_elements_years.first().child_leaf.credits \
            if group_elements_years.first().child_leaf.credits else EMPTY


def _get_volumes(learning_unit_yr):
    volumes = {
        learning_component_year_type.LECTURING: _initialize_component_data(),
        learning_component_year_type.PRACTICAL_EXERCISES: _initialize_component_data()
    }
    data_components = get_same_container_year_components(learning_unit_yr, True)

    for component in data_components.get('components'):
        if component.get('learning_component_year').type in (learning_component_year_type.LECTURING,
                                                             learning_component_year_type.PRACTICAL_EXERCISES):
            volumes = _update_volumes_data(component, volumes)

    return volumes


def _update_volumes_data(compo, volumes):
    vol_to_update = volumes.get(compo.get('learning_component_year').type)
    key_of_value_to_update = ['VOLUME_TOTAL', 'VOLUME_Q1', 'VOLUME_Q2', 'PLANNED_CLASSES']
    for key in key_of_value_to_update:
        if compo.get('volumes').get(key):
            vol_to_update[key] = vol_to_update.get(key) + compo.get('volumes').get(key)

    volumes[compo.get('learning_component_year').type] = vol_to_update
    return volumes


def _get_significant_volume(volume):
    if volume and volume > 0:
        return volume
    return ''


def _initialize_component_data():
    return {
        'VOLUME_TOTAL': 0,
        'PLANNED_CLASSES': 0,
        'VOLUME_Q1': 0,
        'VOLUME_Q2': 0
    }


def _prepare_legend_ws_data():
    return {
        xls_build.HEADER_TITLES_KEY: [str(_('Legend'))],
        xls_build.CONTENT_KEY: [
            ['  ', _('proposal_creation')],
            ['  ', _('Proposal for modification')],
            ['  ', _('Suppression proposal')],
            ['  ', _('Transformation proposal')],
            ['  ', _('Transformation/modification proposal')],
        ],
        xls_build.WORKSHEET_TITLE_KEY: _('Legend'),
        xls_build.STYLED_CELLS:
            {
                Style(fill=PatternFill(patternType='solid', fgColor=CREATION_COLOR)): ['A2'],
                Style(fill=PatternFill(patternType='solid', fgColor=MODIFICATION_COLOR)): ['A3'],
                Style(fill=PatternFill(patternType='solid', fgColor=SUPPRESSION_COLOR)): ['A4'],
                Style(fill=PatternFill(patternType='solid', fgColor=TRANSFORMATION_COLOR)): ['A5'],
                Style(fill=PatternFill(patternType='solid', fgColor=TRANSFORMATION_AND_MODIFICATION)): ['A6'],
            }
    }


def _get_formations_by_educ_group_year(found_learning_units):
    groups = []
    for learning_unit_yr in found_learning_units:
        learning_unit_yr.group_elements_years = mdl_base.group_element_year.search(child_leaf=learning_unit_yr) \
            .select_related("parent", "child_leaf", "parent__education_group_type") \
            .order_by('parent__partial_acronym')
        groups.extend(learning_unit_yr.group_elements_years)
    education_groups_years = [group_element_year.parent for group_element_year in groups]
    return mdl_base.group_element_year \
        .find_learning_unit_formations(education_groups_years, parents_as_instances=True)


def _get_format(found_learning_units):
    dict_wrapped_styled_cells = {}
    dict_colored_styled_cells = {}

    for idx, luy in enumerate(found_learning_units, start=2):
        dict_wrapped_styled_cells.update({WRAP_TEXT_STYLE:  ["X{}".format(idx), "Y{}".format(idx)]})
        proposal = mdl_base.proposal_learning_unit.find_by_learning_unit_year(luy)
        if proposal:
            dict_colored_styled_cells.update({PROPOSAL_LINE_STYLES.get(proposal.type): [idx-1]})
    return {
        WRAP: dict_wrapped_styled_cells,
        COLORED: dict_colored_styled_cells
    }

def _get_attribution_line(ch, value):
    ch = ch + "{} - {} : {} - {} : {} - {} : {} - {} : {} - {} : {} - {} : {} ".format(
        value.get('person'),
        _('function'),
        _(value.get('function')) if value.get('function') else '',
        _('substitute'),
        value.get('substitute') if value.get('substitute') else '',
        _('Beg. of attribution'),
        value.get('start_year'),
        _('Attribution duration'),
        value.get('duration'),
        _('Attrib. vol1'),
        value.get('LECTURING'),
        _('Attrib. vol2'),
        value.get('PRACTICAL_EXERCISES'),
    )
    return ch
