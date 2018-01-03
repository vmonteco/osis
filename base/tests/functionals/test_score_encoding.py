import datetime
import os
import pdb
import random
import time

import faker
import pendulum
from django.contrib.auth.models import Permission, Group
from django.contrib.staticfiles.testing import StaticLiveServerTestCase
from django.urls import reverse
from openpyxl import load_workbook
from prettyprinter import cpprint
from selenium import webdriver

from base.models import session_exam_calendar
from base.tests.factories.academic_calendar import AcademicCalendarFactory, AcademicCalendarExamSubmissionFactory
from base.tests.factories.academic_year import AcademicYearFactory
from base.tests.factories.exam_enrollment import ExamEnrollmentFactory
from base.tests.factories.learning_unit import LearningUnitFactory
from base.tests.factories.learning_unit_enrollment import LearningUnitEnrollmentFactory
from base.tests.factories.learning_unit_year import LearningUnitYearFactory
from base.tests.factories.offer_enrollment import OfferEnrollmentFactory
from base.tests.factories.offer_year import OfferYearFactory
from base.tests.factories.offer_year_calendar import OfferYearCalendarFactory
from base.tests.factories.person import PersonFactory
from base.tests.factories.program_manager import ProgramManagerFactory
from base.tests.factories.session_exam_calendar import SessionExamCalendarFactory
from base.tests.factories.session_examen import SessionExamFactory
from base.tests.factories.student import StudentFactory
from base.tests.factories.user import UserFactory, SuperUserFactory

SIZE = (1280, 1280)

class BusinessMixin:
    def create_user(self):
        return UserFactory()

    def create_super_user(self):
        return SuperUserFactory()

class SeleniumTestCase(StaticLiveServerTestCase):
    def setUp(self):
        import pyvirtualdisplay
        self.display = pyvirtualdisplay.Display(visible=0, size=SIZE)
        # self.display.start()
        options = webdriver.ChromeOptions()
        options.add_experimental_option('prefs', {
            'download.default_directory': '/tmp',
            'download.prompt_for_download': False,
            'download.directory_upgrade': True,
            'safebrowsing.enabled': True
        })
        self.driver = webdriver.Chrome(chrome_options=options)
        self.driver.implicitly_wait(10)
        self.driver.set_window_size(*SIZE)

    def tearDown(self):
        self.driver.quit()
        # self.display.stop()

    def get_url_by_name(self, url_name, *args, **kwargs):
        url = '{}{}'.format(self.live_server_url, reverse(url_name, args=args, kwargs=kwargs))
        return url

    def goto(self, url_name, *args, **kwargs):
        url = self.get_url_by_name(url_name, *args, **kwargs)
        self.driver.get(url)

    def fill_by_id(self, field_id, value):
        field = self.driver.find_element_by_id(field_id)
        field.clear()
        field.send_keys(value)

    def login(self, username, password='password'):
        self.goto('login')
        self.fill_by_id('id_username', username)
        self.fill_by_id('id_password', password)
        self.driver.find_element_by_id('post_login_btn').click()


class ScoreEncodingTestCase(SeleniumTestCase, BusinessMixin):
    def test_01_scenario_modifier_periode_encoding(self):
        user = self.create_super_user()
        permission = Permission.objects.get(codename='can_access_academic_calendar')
        user.user_permissions.add(permission)
        self.login(user.username, 'password123')

        academic_year = AcademicYearFactory()
        academic_calendar = AcademicCalendarFactory.build(academic_year=academic_year)
        academic_calendar.save(functions=[])

        self.goto('academic_calendar_read', academic_calendar_id=academic_calendar.id)
        self.driver.find_element_by_id('bt_academic_calendar_edit').click()

        # pdb.set_trace()

        new_date = academic_calendar.start_date - datetime.timedelta(days=5)
        new_date_str = new_date.strftime('%m/%d/%Y')
        self.fill_by_id('txt_start_date', new_date_str)

        self.driver.save_screenshot('/tmp/screenshot.png')

        self.driver.execute_script("scroll(0, 250)")
        self.driver.find_element_by_id('bt_academic_calendar_save').click()

        self.assertEqual(
            self.driver.current_url,
            self.get_url_by_name('academic_calendar_form', academic_calendar_id=academic_calendar.id)
        )

        self.assertEqual(self.driver.find_element_by_id('ac_start_date').text, new_date_str)

    def test_01_scenario_modifier_period_encoding_date_fin(self):
        user = self.create_super_user()
        permission = Permission.objects.get(codename='can_access_academic_calendar')
        user.user_permissions.add(permission)
        self.login(user.username, 'password123')

        academic_year = AcademicYearFactory()
        academic_calendar = AcademicCalendarFactory.build(academic_year=academic_year)
        academic_calendar.save(functions=[])

        self.goto('academic_calendar_read', academic_calendar_id=academic_calendar.id)
        self.driver.find_element_by_id('bt_academic_calendar_edit').click()

        # pdb.set_trace()

        new_date = academic_calendar.end_date + datetime.timedelta(days=5)
        new_date_str = new_date.strftime('%m/%d/%Y')
        self.fill_by_id('txt_end_date', new_date_str)

        self.driver.save_screenshot('/tmp/screenshot2.png')

        self.driver.execute_script("scroll(0, 250)")
        self.driver.find_element_by_id('bt_academic_calendar_save').click()

        self.assertEqual(
            self.driver.current_url,
            self.get_url_by_name('academic_calendar_form', academic_calendar_id=academic_calendar.id)
        )

        self.assertEqual(self.driver.find_element_by_id('ac_end_date').text, new_date_str)


class TestScoreEncodingScenario4(SeleniumTestCase, BusinessMixin):
    def test_04(self):
        user = self.create_user()  # create_super_user()
        group = Group.objects.create(name='program_managers')
        group2 = Group.objects.create(name='tutors')

        permission = Permission.objects.get(codename='can_access_academic_calendar')
        user.user_permissions.add(permission)
        permission = Permission.objects.filter(codename='can_access_scoreencoding',
                                               content_type__app_label='assessments').first()

        user.user_permissions.add(permission)

        self.login(user.username, 'password123')

        academic_year = AcademicYearFactory(year=pendulum.today().year)
        academic_calendar = AcademicCalendarExamSubmissionFactory.build(academic_year=academic_year)
        academic_calendar.save(functions=[])
        # cpprint({k: getattr(academic_calendar, k) for k in ('start_date', 'end_date', 'reference')})

        person = PersonFactory(user=user)
        offer_year_phys11ba = OfferYearFactory(academic_year=academic_year, acronym='PHYS11BA')
        offer_year_econ2m1 = OfferYearFactory(academic_year=academic_year, acronym='ECON2M1')
        offer_year_phys1ba = OfferYearFactory(academic_year=academic_year, acronym='PHYS1BA')
        offer_year_phys2m1 = OfferYearFactory(academic_year=academic_year, acronym='PHYS2M1')
        offer_year_phys2ma = OfferYearFactory(academic_year=academic_year, acronym='PHYS2MA')

        ProgramManagerFactory(offer_year=offer_year_phys11ba, person=person)
        ProgramManagerFactory(offer_year=offer_year_econ2m1, person=person)
        ProgramManagerFactory(offer_year=offer_year_phys1ba, person=person)
        ProgramManagerFactory(offer_year=offer_year_phys2m1, person=person)
        ProgramManagerFactory(offer_year=offer_year_phys2ma, person=person)

        student1 = StudentFactory()
        student2 = StudentFactory()
        student3 = StudentFactory()
        student4 = StudentFactory()
        student5 = StudentFactory()

        learning_unit = LearningUnitFactory()
        learning_unit_year = LearningUnitYearFactory(academic_year=academic_year, learning_unit=learning_unit)

        # learning_unit2 = LearningUnitFactory()
        # learning_unit_year2 = LearningUnitYearFactory(academic_year=academic_year, learning_unit=learning_unit2)

        offer_enrollment1 = OfferEnrollmentFactory(offer_year=offer_year_phys11ba, student=student1)
        offer_enrollment2 = OfferEnrollmentFactory(offer_year=offer_year_phys11ba, student=student2)
        offer_enrollment3 = OfferEnrollmentFactory(offer_year=offer_year_phys11ba, student=student3)
        offer_enrollment4 = OfferEnrollmentFactory(offer_year=offer_year_phys11ba, student=student4)
        offer_enrollment5 = OfferEnrollmentFactory(offer_year=offer_year_phys11ba, student=student5)

        learning_unit_enrollment1 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment1,
                                                                  learning_unit_year=learning_unit_year)
        learning_unit_enrollment2 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment2,
                                                                  learning_unit_year=learning_unit_year)
        learning_unit_enrollment3 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment3,
                                                                  learning_unit_year=learning_unit_year)
        learning_unit_enrollment4 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment4,
                                                                  learning_unit_year=learning_unit_year)
        learning_unit_enrollment5 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment5,
                                                                  learning_unit_year=learning_unit_year)

        number_session = 1
        session_exam = SessionExamFactory(learning_unit_year=learning_unit_year, number_session=number_session,
                                          offer_year=offer_year_phys11ba)

        # session_exam2 = SessionExamFactory(learning_unit_year=learning_unit_year2, number_session=number_session, offer_year=offer_year_phys11ba)

        offer_year_calendar = OfferYearCalendarFactory(academic_calendar=academic_calendar, offer_year=offer_year_phys11ba)

        enrollments = []
        for counter in range(1, 6):
            ExamEnrollmentFactory(learning_unit_enrollment=locals()['learning_unit_enrollment{}'.format(counter)],
                                  session_exam=session_exam)

        # ExamEnrollmentFactory(learning_unit_enrollment1)
        # FIXME: 11 unités d'enseignement dans 1 programme(s).
        # ok, mais d'ou vient ce 11 unité d'enseignement? comment le calculer, probleme de vocabulaire

        sec = SessionExamCalendarFactory(academic_calendar=academic_calendar)

        self.goto('scores_encoding')

        from selenium.webdriver.support.ui import Select

        select = Select(self.driver.find_element_by_id('slt_offer_list_selection'))
        select.select_by_value(str(offer_year_phys11ba.id))

        self.driver.find_element_by_id('bt_submit_offer_search').click()

        number_of_learning_units = int(self.driver.find_element_by_id('scores_encoding_learning_units').text.strip())
        number_of_programs = int(self.driver.find_element_by_id('scores_encoding_programs').text.strip())

        print(number_of_learning_units, number_of_programs)

        self.assertEqual(number_of_programs, 1)
        self.assertEqual(number_of_learning_units, 1)

        # pdb.set_trace()


        #scores_encoding_programs
        # self.driver.find_element_by_id('lnk_via_excel').click()
        # time.sleep(1)
        # self.driver.find_element_by_id('lnk_scores_encoding_download_{}'.format(learning_unit_year.id)).click()
        # # time.sleep(1)
        # filename = 'session_{}_{}_{}.xlsx'.format(academic_year.year, number_session, learning_unit_year.acronym)
        # full_path = os.path.join('/', 'tmp', filename)
        #
        # print(full_path)
        # self.assertTrue(os.path.exists(full_path))
        #
        # update_xslx(full_path, enrollments)
        # # pdb.set_trace()
        #
        # self.goto('online_encoding', learning_unit_year_id=learning_unit_year.id)
        # self.driver.find_element_by_id('bt_upload_score_modal').click()
        # time.sleep(1)
        # self.driver.execute_script("document.getElementById('fle_scores_input_file').style.display = 'block'")
        # self.fill_by_id('fle_scores_input_file', full_path)
        # # time.sleep(1)
        # self.driver.find_element_by_id('bt_submit_upload_score_modal').click()
        #
        # progression = self.driver.find_element_by_id('luy_progression').text
        # self.assertEqual(progression, '0 / 5')

class TestScoreEncodingWithoutSeleniumTransition(SeleniumTestCase, BusinessMixin):
    def test_05_score_encoding(self):
        # assessments.views.score_encoding.scores_encoding
        user = self.create_user() # create_super_user()
        group = Group.objects.get_or_create(name='program_managers')
        # group = Group.objects.create(name='program_managers')
        group2 = Group.objects.get_or_create(name='tutors')
        # group2 = Group.objects.create(name='tutors')

        permission = Permission.objects.get(codename='can_access_academic_calendar')
        user.user_permissions.add(permission)
        permission = Permission.objects.filter(codename='can_access_scoreencoding', content_type__app_label='assessments').first()

        user.user_permissions.add(permission)

        self.login(user.username, 'password123')

        academic_year = AcademicYearFactory(year=pendulum.today().year)
        academic_calendar = AcademicCalendarExamSubmissionFactory.build(academic_year=academic_year)
        academic_calendar.save(functions=[])
        # cpprint({k: getattr(academic_calendar, k) for k in ('start_date', 'end_date', 'reference')})

        person = PersonFactory(user=user)
        offer_year_phys11ba = OfferYearFactory(academic_year=academic_year, acronym='PHYS11BA')
        offer_year_econ2m1 = OfferYearFactory(academic_year=academic_year, acronym='ECON2M1')
        offer_year_phys1ba = OfferYearFactory(academic_year=academic_year, acronym='PHYS1BA')
        offer_year_phys2m1 = OfferYearFactory(academic_year=academic_year, acronym='PHYS2M1')
        offer_year_phys2ma = OfferYearFactory(academic_year=academic_year, acronym='PHYS2MA')

        ProgramManagerFactory(offer_year=offer_year_phys11ba, person=person)
        ProgramManagerFactory(offer_year=offer_year_econ2m1, person=person)
        ProgramManagerFactory(offer_year=offer_year_phys1ba, person=person)
        ProgramManagerFactory(offer_year=offer_year_phys2m1, person=person)
        ProgramManagerFactory(offer_year=offer_year_phys2ma, person=person)

        student1 = StudentFactory()
        student2 = StudentFactory()
        student3 = StudentFactory()
        student4 = StudentFactory()
        student5 = StudentFactory()

        learning_unit = LearningUnitFactory()
        learning_unit_year = LearningUnitYearFactory(academic_year=academic_year, learning_unit=learning_unit)

        # tutor = TutorFactory()
        # AttributionFactory(learning_unit_year=learning_unit_year, tutor=tutor, score_responsible=True)

        offer_enrollment1 = OfferEnrollmentFactory(offer_year=offer_year_phys11ba, student=student1)
        offer_enrollment2 = OfferEnrollmentFactory(offer_year=offer_year_phys11ba, student=student2)
        offer_enrollment3 = OfferEnrollmentFactory(offer_year=offer_year_phys11ba, student=student3)
        offer_enrollment4 = OfferEnrollmentFactory(offer_year=offer_year_phys11ba, student=student4)
        offer_enrollment5 = OfferEnrollmentFactory(offer_year=offer_year_phys11ba, student=student5)

        learning_unit_enrollment1 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment1, learning_unit_year=learning_unit_year)
        learning_unit_enrollment2 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment2, learning_unit_year=learning_unit_year)
        learning_unit_enrollment3 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment3, learning_unit_year=learning_unit_year)
        learning_unit_enrollment4 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment4, learning_unit_year=learning_unit_year)
        learning_unit_enrollment5 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment5, learning_unit_year=learning_unit_year)


        number_session = 1
        session_exam = SessionExamFactory(learning_unit_year=learning_unit_year, number_session=number_session, offer_year=offer_year_phys11ba)
        offer_year_calendar = OfferYearCalendarFactory(academic_calendar=academic_calendar, offer_year=offer_year_phys11ba)

        enrollments = []
        for counter in range(1, 6):
            enrollments.append({
                'enrollment': ExamEnrollmentFactory(
                    learning_unit_enrollment=locals()['learning_unit_enrollment{}'.format(counter)],
                    session_exam=session_exam
                ),
                'note': None,
                'justification': None
            })

        sec = SessionExamCalendarFactory(academic_calendar=academic_calendar)

        diff_date = academic_calendar.end_date - academic_calendar.start_date

        days = diff_date.days / 2
        date = academic_calendar.start_date + datetime.timedelta(days=days)
        # cpprint({'diff_date': diff_date, 'days': days, 'date': date})

        self.assertIsNotNone(session_exam_calendar.current_session_exam(date))

        # Il faut une periode ouverte, sinon, on est retourne vers
        # assessments.views.score_encoding.outside_period

        self.goto('scores_encoding')
        # pdb.set_trace()
        self.driver.find_element_by_id('lnk_via_excel').click()
        time.sleep(1)
        self.driver.find_element_by_id('lnk_scores_encoding_download_{}'.format(learning_unit_year.id)).click()
        # time.sleep(1)
        filename = 'session_{}_{}_{}.xlsx'.format(academic_year.year, number_session, learning_unit_year.acronym)
        full_path = os.path.join('/', 'tmp', filename)

        print(full_path)
        self.assertTrue(os.path.exists(full_path))

        update_xslx(full_path, enrollments)
        # pdb.set_trace()

        self.goto('online_encoding', learning_unit_year_id=learning_unit_year.id)
        self.driver.find_element_by_id('bt_upload_score_modal').click()
        time.sleep(1)
        self.driver.execute_script("document.getElementById('fle_scores_input_file').style.display = 'block'")
        self.fill_by_id('fle_scores_input_file', full_path)
        # time.sleep(1)
        self.driver.find_element_by_id('bt_submit_upload_score_modal').click()

        progression = self.driver.find_element_by_id('luy_progression').text
        self.assertEqual(progression, '0 / 5')

        for enrollment_idx, value in enumerate(enrollments):
            enrollment = value['enrollment']
            note = self.driver.find_element_by_id('enrollment_note_{}'.format(enrollment.id)).text
            if note != '-' and value['note'] is not None:
                self.assertEqual(value['note'], int(note))
            justification = self.driver.find_element_by_id('enrollment_justification_{}'.format(enrollment.id)).text
            if note != '-' and value['justification'] is not None:
                self.assertEqual(value['justification'], justification)

        # pdb.set_trace()
        # from openpyxl import load_workbook

        # Scenario 5, il y a un probleme, car je n'ai pas les memes registration_id dans les etudiants, et donc, l'algo ne les retrouve pas
        # ce qui fait que sans ce petit correctif, le test est erroné.


class TestPrintPdf(SeleniumTestCase, BusinessMixin):
    def test_06_print_pdf(self):
        user = self.create_user() # create_super_user()
        group = Group.objects.get_or_create(name='program_managers')
        # group = Group.objects.create(name='program_managers')
        # group2 = Group.objects.create(name='tutors')
        group2 = Group.objects.get_or_create(name='tutors')

        permission = Permission.objects.get(codename='can_access_academic_calendar')
        user.user_permissions.add(permission)
        permission = Permission.objects.filter(codename='can_access_scoreencoding', content_type__app_label='assessments').first()

        user.user_permissions.add(permission)

        self.login(user.username, 'password123')

        academic_year = AcademicYearFactory(year=pendulum.today().year)
        academic_calendar = AcademicCalendarExamSubmissionFactory.build(academic_year=academic_year)
        academic_calendar.save(functions=[])
        # cpprint({k: getattr(academic_calendar, k) for k in ('start_date', 'end_date', 'reference')})

        person = PersonFactory(user=user)
        offer_year_phys11ba = OfferYearFactory(academic_year=academic_year, acronym='PHYS11BA')
        offer_year_econ2m1 = OfferYearFactory(academic_year=academic_year, acronym='ECON2M1')
        offer_year_phys1ba = OfferYearFactory(academic_year=academic_year, acronym='PHYS1BA')
        offer_year_phys2m1 = OfferYearFactory(academic_year=academic_year, acronym='PHYS2M1')
        offer_year_phys2ma = OfferYearFactory(academic_year=academic_year, acronym='PHYS2MA')

        ProgramManagerFactory(offer_year=offer_year_phys11ba, person=person)
        ProgramManagerFactory(offer_year=offer_year_econ2m1, person=person)
        ProgramManagerFactory(offer_year=offer_year_phys1ba, person=person)
        ProgramManagerFactory(offer_year=offer_year_phys2m1, person=person)
        ProgramManagerFactory(offer_year=offer_year_phys2ma, person=person)

        student1 = StudentFactory()
        student2 = StudentFactory()
        student3 = StudentFactory()
        student4 = StudentFactory()
        student5 = StudentFactory()

        learning_unit = LearningUnitFactory()
        learning_unit_year = LearningUnitYearFactory(academic_year=academic_year, learning_unit=learning_unit)

        # tutor = TutorFactory()
        # AttributionFactory(learning_unit_year=learning_unit_year, tutor=tutor, score_responsible=True)

        offer_enrollment1 = OfferEnrollmentFactory(offer_year=offer_year_phys11ba, student=student1)
        offer_enrollment2 = OfferEnrollmentFactory(offer_year=offer_year_phys11ba, student=student2)
        offer_enrollment3 = OfferEnrollmentFactory(offer_year=offer_year_phys11ba, student=student3)
        offer_enrollment4 = OfferEnrollmentFactory(offer_year=offer_year_phys11ba, student=student4)
        offer_enrollment5 = OfferEnrollmentFactory(offer_year=offer_year_phys11ba, student=student5)

        learning_unit_enrollment1 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment1, learning_unit_year=learning_unit_year)
        learning_unit_enrollment2 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment2, learning_unit_year=learning_unit_year)
        learning_unit_enrollment3 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment3, learning_unit_year=learning_unit_year)
        learning_unit_enrollment4 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment4, learning_unit_year=learning_unit_year)
        learning_unit_enrollment5 = LearningUnitEnrollmentFactory(offer_enrollment=offer_enrollment5, learning_unit_year=learning_unit_year)


        number_session = 1
        session_exam = SessionExamFactory(learning_unit_year=learning_unit_year, number_session=number_session, offer_year=offer_year_phys11ba)
        offer_year_calendar = OfferYearCalendarFactory(academic_calendar=academic_calendar, offer_year=offer_year_phys11ba)

        enrollments = []
        for counter in range(1, 6):
            enrollments.append({
                'enrollment': ExamEnrollmentFactory(
                    learning_unit_enrollment=locals()['learning_unit_enrollment{}'.format(counter)],
                    session_exam=session_exam
                ),
                'note': None,
                'justification': None
            })

        sec = SessionExamCalendarFactory(academic_calendar=academic_calendar)

        self.goto('scores_encoding')
        self.fill_by_id('txt_acronym', learning_unit_year.acronym)
        self.driver.find_element_by_id('bt_submit_offer_search').click()
        time.sleep(2)

        self.driver.find_element_by_id('lnk_via_paper').click()
        time.sleep(1)

        self.driver.find_element_by_id('lnk_notes_printing_{}'.format(learning_unit_year.id)).click()
        time.sleep(2)
        filename = 'Feuille de notes.pdf'
        full_path = os.path.join('/', 'tmp', filename)

        self.assertTrue(os.path.exists(full_path))
        import magic
        mimetype = magic.from_file(full_path, mime=True)
        self.assertEqual(mimetype, 'application/pdf')


def update_xslx(filename, enrollments):
    fake = faker.Faker()

    wb = load_workbook(filename)

    sheet = wb.active

    if sheet.max_row > 11:
        start_row = 12
        for row_number, enrollment in enumerate(enrollments): # range(12, sheet.max_row + 1):
            left_or_right = bool(random.getrandbits(1))
            selected_column = 'H' if left_or_right else 'I'

            if left_or_right:
                value = random.randint(0, 20)
                key = 'note'
            else:
                value = fake.random_element(elements=('A', 'T', 'S', 'M'))
                key = 'justification'

            sheet['{}{}'.format(selected_column, row_number+1)] = value
            enrollment[key] = value

    wb.save(filename=filename)
