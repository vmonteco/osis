import datetime
import os
import random
import time

import faker
import magic
import parse
import pendulum
from behave import given, when, then
from django.contrib.auth.models import Group, Permission
from django.utils import timezone
from openpyxl import load_workbook
from selenium.webdriver.support.select import Select

from base.tests.factories.academic_calendar import AcademicCalendarExamSubmissionFactory
from base.tests.factories.academic_year import AcademicYearFactory
from base.tests.factories.exam_enrollment import ExamEnrollmentFactory
from base.tests.factories.learning_unit_enrollment import LearningUnitEnrollmentFactory
from base.tests.factories.learning_unit_year import LearningUnitYearFactory
from base.tests.factories.offer_enrollment import OfferEnrollmentFactory
from base.tests.factories.offer_year import OfferYearFactory
from base.tests.factories.offer_year_calendar import OfferYearCalendarFactory
from base.tests.factories.person import PersonFactory
from base.tests.factories.program_manager import ProgramManagerFactory
from base.tests.factories.session_exam_calendar import SessionExamCalendarFactory
from base.tests.factories.session_examen import SessionExamFactory
from base.tests.factories.student import StudentFactory
from base.tests.factories.user import SuperUserFactory, UserFactory

fake = faker.Faker()


@given('I am a Super User')
def step_impl(context):
    context.user = SuperUserFactory()
    context.person = PersonFactory(user=context.user, language='fr-be')


@given('I am a User')
def step_impl(context):
    context.user = UserFactory()
    context.person = PersonFactory(user=context.user, language='fr-be')


@given('I am a Program Manager')
def step_impl(context):
    context.execute_steps("""
        Given I am a User
    """)

    group_names = ('program_managers',)
    for name in group_names:
        group, created = Group.objects.get_or_create(name=name)
        group.user_set.add(context.user)

    permission_names = (
        'can_access_academic_calendar',
        'assessments.can_access_scoreencoding'
    )

    for permission_name in permission_names:
        if '.' in permission_name:
            label, codename = permission_name.split('.')
            permission = Permission.objects.get(codename=codename, content_type__app_label=label)
        else:
            permission = Permission.objects.get(codename=permission_name)

        context.user.user_permissions.add(permission)


@given('I am the Program Manager of this Academic Calendar')
def step_impl(context):
    context.offer_year = OfferYearFactory(academic_year=context.academic_year)
    context.program_manager = ProgramManagerFactory(offer_year=context.offer_year, person=context.person)


@given('I am logging in')
def step_impl(context):
    context.browser.login(context.user.username)


@given('There is an Academic Calendar')
def step_impl(context):
    academic_year = AcademicYearFactory(year=pendulum.today().year - 1)
    academic_calendar = AcademicCalendarExamSubmissionFactory.build(academic_year=academic_year)
    academic_calendar.save(functions=[])

    context.academic_year = academic_year
    context.academic_calendar = academic_calendar


@given('I am on the Academic Calendar page')
def step_impl(context):
    context.browser.goto('academic_calendar_read', academic_calendar_id=context.academic_calendar.id)


@when('I change the start date in the form')
def step_impl(context):
    context.browser.click_on('bt_academic_calendar_edit')

    context.date = context.academic_calendar.start_date - datetime.timedelta(days=random.randint(5, 10))
    context.browser.fill_by_id('txt_start_date', context.date.strftime('%m/%d/%Y'))

    context.browser.driver.execute_script("scroll(0, 250)")
    context.browser.click_on('bt_academic_calendar_save')


@when('I change the end date in the form')
def step_impl(context):
    context.browser.click_on('bt_academic_calendar_edit')

    context.date = context.academic_calendar.end_date + datetime.timedelta(days=random.randint(5, 10))
    context.browser.fill_by_id('txt_end_date', context.date.strftime('%m/%d/%Y'))

    context.browser.driver.execute_script("scroll(0, 250)")
    context.browser.click_on('bt_academic_calendar_save')


@then('the start date on the detail view should be equal')
def step_impl(context):
    assert context.browser.driver.current_url == context.get_url('academic_calendar_form',
                                                                 academic_calendar_id=context.academic_calendar.id)
    assert context.browser.get_element_text('ac_start_date') == context.date.strftime('%m/%d/%Y')


@then('the end date on the detail view should be equal')
def step_impl(context):
    assert context.browser.driver.current_url == context.get_url('academic_calendar_form',
                                                                 academic_calendar_id=context.academic_calendar.id)
    assert context.browser.get_element_text('ac_end_date') == context.date.strftime('%m/%d/%Y')


@given('There is an Academic Calendar in the future')
def step_impl(context):
    start_date = timezone.now() + datetime.timedelta(days=20)

    academic_year = AcademicYearFactory(year=pendulum.today().year - 1)

    academic_calendar = AcademicCalendarExamSubmissionFactory.build(
        academic_year=academic_year,
        start_date=start_date,
        end_date=start_date + datetime.timedelta(days=10),
    )
    academic_calendar.save(functions=[])

    context.academic_year = academic_year
    context.academic_calendar = academic_calendar


@given('There is a Session Exam')
def step_impl(context):
    context.session_exam = SessionExamCalendarFactory(academic_calendar=context.academic_calendar)


@when('I am on the Scores Encoding page')
def step_impl(context):
    context.browser.goto('scores_encoding')


@then("the scores encoding period will be open in the future")
def step_impl(context):
    warning_message = context.browser.get_element_text('pnl_warning_messages')

    def parse_date(text):
        return datetime.datetime.strptime(text, '%d/%m/%Y').date()

    result = parse.parse(
        "La période d'encodage des notes pour la session {session:d} sera ouverte à partir du {date:Date}",
        warning_message,
        dict(Date=parse_date)
    )

    assert result is not None
    assert result['date'] == context.academic_calendar.start_date.date()


@given('There are {number:d} Learning Unit Year')
def step_impl(context, number):
    context.learning_unit_years = {}
    for counter in range(number+1):
        context.learning_unit_years[counter+1] = LearningUnitYearFactory(academic_year=context.academic_year)


@given('There are several Offer Years where I am the Program Manager')
def step_impl(context):
    context.offer_years = {}
    for row in context.table:
        acronym = row['acronym']

        offer_year = OfferYearFactory(academic_year=context.academic_year,
                                      acronym=acronym)

        ProgramManagerFactory(person=context.person, offer_year=offer_year)

        context.offer_years[acronym] = offer_year


@given('There are {number:d} students for "{offer_year}" and the Learning Unit Year {learning_unit_year:d}')
def step_impl(context, number, offer_year, learning_unit_year):
    context.learning_unit_enrollments = []
    for counter in range(number):
        student = StudentFactory()
        offer_enrollment = OfferEnrollmentFactory(offer_year=context.offer_years[offer_year],
                                                  student=student)
        learning_unit_enrollment = LearningUnitEnrollmentFactory(
            offer_enrollment=offer_enrollment,
            learning_unit_year=context.learning_unit_years[learning_unit_year]
        )
        context.learning_unit_enrollments.append(learning_unit_enrollment)


@given('There is a Session Exam for "{offer_year}" and the Learning Unit Year {learning_unit_year:d}')
def step_impl(context, offer_year, learning_unit_year):

    session_exam_calendar = SessionExamCalendarFactory(academic_calendar=context.academic_calendar)
    context.session_exam_calendar = session_exam_calendar

    session_exam = SessionExamFactory(
        learning_unit_year=context.learning_unit_years[learning_unit_year],
        number_session=session_exam_calendar.number_session,
        offer_year=context.offer_years[offer_year]
    )
    context.session_exam = session_exam

    OfferYearCalendarFactory(
        academic_calendar=context.academic_calendar,
        offer_year=context.offer_years[offer_year]
    )

    context.exam_enrollments = []
    for learning_unit_enrollment in context.learning_unit_enrollments:
        exam_enrollment = ExamEnrollmentFactory(
            learning_unit_enrollment=learning_unit_enrollment,
            session_exam=session_exam
        )
        context.exam_enrollments.append(exam_enrollment)


@then('I have the previous Offer Years in the list of Programs')
def step_impl(context):
    select = Select(context.browser.get_element('slt_offer_list_selection'))
    all_options = set(option.text for option in select.options)
    all_offers = set(context.offer_years.keys())

    assert {'Tout'} == (all_options - all_offers)


@then('There is 1 Learning Unit')
@then('There are {number:d} Learning Units')
def step_impl(context, number=1):
    number_of_learning_units = context.browser.get_element_int('scores_encoding_learning_units')
    assert number_of_learning_units == number


@when('I want to encode the Learning Unit Year {learning_unit_year:d}')
def step_impl(context, learning_unit_year):
    learning_unit_year = context.learning_unit_years[learning_unit_year]
    context.browser.click_on('lnk_encode_{}'.format(learning_unit_year.id))
    context.current_learning_unit_year = learning_unit_year


@then('The decimal scores are not possible')
def step_impl(context):
    message = context.browser.get_element_text('message_decimal_accepted')
    assert 'Les notes de ce cours ne peuvent PAS' in message


@then('The number of enrollments is {expected_number:d}')
def step_impl(context, expected_number):
    number_of_enrollment = context.browser.get_element_int('number_of_enrollments')
    assert number_of_enrollment == expected_number


@when('I change the note of the first enrollment with {note:d}')
def step_impl(context, note):
    the_first = 1
    element = context.browser.driver.find_element_by_css_selector("[tabindex='{}']".format(the_first))
    element_id = element.get_attribute('id')
    element.clear()
    element.send_keys(note)

    enrollment_id = int(element_id.split('_')[-1])
    context.current_enrollment_id = enrollment_id


@when('I save the online encoding')
def step_impl(context):
    context.browser.click_on('bt_save_online_encoding_up')


@then('The progression must be {number_enrollments_with_status:d} on {maximum_enrollments:d}')
def step_impl(context, number_enrollments_with_status, maximum_enrollments):
    progression = context.browser.get_element_text('luy_progression')
    assert progression == '{} / {}'.format(number_enrollments_with_status, maximum_enrollments)


@then('The note for this enrollment is {expected_note:d}')
def step_impl(context, expected_note):
    note = context.browser.get_element_int('enrollment_note_{}'.format(context.current_enrollment_id))
    assert note == expected_note


@then('This enrollment has the plane icon')
def step_impl(context):
    css_selector = 'td#enrollment_status_{} span'.format(context.current_enrollment_id)
    element = context.browser.driver.find_element_by_css_selector(css_selector)
    assert 'glyphicon-send' in element.get_attribute('class').split()


@when('I change the notes of the enrollments')
def step_impl(context):

    computed_values = compute_notes_and_justifications(context.exam_enrollments)
    for enrollment_id, (key, value) in computed_values.items():
        if key == 'note':
            element_id = 'num_score_{}'.format(enrollment_id)
            context.browser.fill_by_id(element_id, value)
        else:
            element_id = 'slt_justification_score_{}'.format(enrollment_id)
            select = Select(context.browser.get_element(element_id))
            value = {'T': 'Tricherie', 'A': 'Absence injustifiée'}.get(value)
            select.select_by_visible_text(value)

    context.note_enrollments = computed_values


@then('The enrollments have the computed values')
def step_impl(context):
    for enrollment_id, (key, expected_value) in context.note_enrollments.items():
        element_id = 'enrollment_{}_{}'.format(key, enrollment_id)
        expected_value = {'T': 'Tricherie', 'A': 'Absence injustifiée'}.get(expected_value, expected_value)

        value = str(context.browser.get_element_text(element_id))
        assert value == str(expected_value)


@then('The enrollments have the plane icon')
def step_impl(context):
    for enrollment_id in context.note_enrollments.keys():
        css_selector = 'td#enrollment_status_{} span'.format(enrollment_id)
        element = context.browser.driver.find_element_by_css_selector(css_selector)
        assert 'glyphicon-send' in element.get_attribute('class').split()


@when('I search for the "{offer_year}" offer')
def step_impl(context, offer_year):
    select = Select(context.browser.get_element('slt_offer_list_selection'))
    select.select_by_visible_text(offer_year.upper())
    context.browser.click_on('bt_submit_offer_search')
    time.sleep(0.5)


@when('I download the Excel file of Learning Unit Year {learning_unit_year_number:d}')
def step_impl(context, learning_unit_year_number):
    context.browser.click_on('lnk_via_excel')
    time.sleep(1)
    learning_unit_year = context.learning_unit_years[learning_unit_year_number]
    element_id = 'lnk_scores_encoding_download_{}'.format(learning_unit_year.id)
    context.browser.click_on(element_id)
    time.sleep(1)

    filename = 'session_{}_{}_{}.xlsx'.format(context.academic_year.year,
                                              context.session_exam_calendar.number_session,
                                              context.learning_unit_years[learning_unit_year].acronym)
    full_path = os.path.join(context.full_path_temp_dir, filename)
    assert os.path.exists(full_path)

    context.full_path_xls = full_path


def compute_notes_and_justifications(exam_enrollments):
    enrollments = {}

    for enrollment in exam_enrollments:
        left_or_right = bool(random.getrandbits(1))

        if left_or_right:
            value = random.randint(0, 20)
            key = 'note'
        else:
            value = fake.random_element(elements=('A', 'T'))
            key = 'justification'

        enrollments[enrollment.id] = (key, value)

    return enrollments


def update_xlsx(filename, exam_enrollments):

    wb = load_workbook(filename)

    enrollments = {}

    sheet = wb.active

    if sheet.max_row > 11:
        start_row = 12

        nomas = {
            enrollment.learning_unit_enrollment.offer_enrollment.student.registration_id: {
                'enrollment': enrollment,
                'position': None
            }
            for enrollment in exam_enrollments
        }

        for counter in range(len(exam_enrollments)):
            noma = sheet['E{}'.format(counter + start_row)].value
            nomas[noma]['position'] = counter

        for noma, info in nomas.items():
            left_or_right = bool(random.getrandbits(1))
            selected_column = 'H' if left_or_right else 'I'

            if left_or_right:
                value = random.randint(0, 20)
                key = 'note'
            else:
                value = fake.random_element(elements=('A', 'T'))
                key = 'justification'

            sheet['{}{}'.format(selected_column, info['position'] + start_row)] = value

            enrollments[info['enrollment'].id] = key, value

    wb.save(filename=filename)
    return enrollments


@when('I upload a modified version of the Excel file')
def step_impl(context):
    context.note_enrollments = update_xlsx(context.full_path_xls, context.exam_enrollments)
    context.browser.goto('online_encoding', learning_unit_year_id=context.learning_unit_years[1].id)
    context.browser.click_on('bt_upload_score_modal')
    time.sleep(0.5)
    context.browser.driver.execute_script("document.getElementById('fle_scores_input_file').style.display = 'block'")
    context.browser.fill_by_id('fle_scores_input_file', context.full_path_xls)
    context.browser.click_on('bt_submit_upload_score_modal')


@when('I search for the "{acronym}" acronym')
def step_impl(context, acronym):
    context.browser.fill_by_id('txt_acronym', acronym)
    context.browser.click_on('bt_submit_offer_search')


@then('I load the PDF file of the Learning Unit Year {learning_unit_year:d}')
def step_impl(context, learning_unit_year):
    context.browser.click_on('lnk_via_paper')
    time.sleep(1)

    link_id = 'lnk_notes_printing_{}'.format(context.learning_unit_years[learning_unit_year].id)
    context.browser.click_on(link_id)

    time.sleep(1)

    filename = 'Feuille de notes.pdf'
    full_path = os.path.join(context.full_path_temp_dir, filename)

    assert os.path.exists(full_path)

    mimetype = magic.from_file(full_path, mime=True)
    assert mimetype == 'application/pdf'


@when('I use the double encoding')
def step_impl(context):
    context.browser.click_on('lnk_online_double_encoding')
    context.double_encoding_points = {}


@when('I force the notes of the enrollments')
def step_impl(context):
    for exam_enrollment_id, (key, value) in context.note_enrollments.items():
        if key == 'note':
            element_id = 'num_double_score_{}'.format(exam_enrollment_id)
            context.browser.fill_by_id(element_id, str(value + 2))
        else:
            element_id = 'slt_double_justification_score_{}'.format(exam_enrollment_id)
            select = Select(context.browser.get_element(element_id))
            value = {'A': 'Tricherie', 'T': 'Absence injustifiée'}.get(value)
            select.select_by_visible_text(value)

    context.browser.click_on('bt_compare_up')
    assert context.browser.driver.current_url == context.get_url('')

    context.browser.driver.execute_script('scroll(0, document.body.scrollHeight)')

    for exam_enrollment_id in context.note_enrollments:
        element_id = 'bt_take_reencoded_{}'.format(exam_enrollment_id)
        context.browser.click_on(element_id)

    context.browser.driver.execute_script('scroll(0, document.body.scrollHeight)')
    context.browser.click_on('bt_submit_online_double_encoding_validation')


@then('The enrollments have the forced values')
def step_impl(context):
    for enrollment_id, (key, expected_value) in context.note_enrollments.items():
        value = context.browser.get_element_text('enrollment_note_{}'.format(enrollment_id))
        if key == 'note':
            assert value == expected_value + 2
